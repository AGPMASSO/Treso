import calendar
import json
import os
import re
import sqlite3
import unicodedata
import urllib.parse
from pathlib import Path
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Optional

import altair as alt
import pandas as pd
import streamlit as st

from persistence import (
    PersistingConnection,
    bootstrap_local_database,
    clear_bootstrap_cache,
    get_persistence_config,
    is_ephemeral_streamlit_host,
    local_db_path,
    persistence_is_configured,
    push_db_to_cloud,
    restore_database_file,
)

LOGO_PATH = Path(__file__).with_name("logo_agpm.png")

DEFAULT_MONTHLY_CONTRIBUTION = 10.0

MONTHS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

PAYMENT_METHODS = ["Espèces", "Virement"]
DEFAULT_PAYMENT_METHOD = PAYMENT_METHODS[0]


def member_ref(member_id: int) -> str:
    return f"M{int(member_id):03d}"


def fmt_v(value: object) -> str:
    """Formate une valeur pour les libellés d'activité ; renvoie '-' si vide/None."""
    if value is None:
        return "-"
    s = str(value).strip()
    return s if s else "-"


def fmt_member_compact(
    reference: str = "",
    nom: str = "",
    prenom: str = "",
    telephone: str = "",
    village: str = "",
    prefecture: str = "",
    email: str = "",
    adresse: str = "",
    date_inscription: str = "",
) -> str:
    """Représentation compacte d'un membre, sans labels (ref nom prenom tel village pref email adresse date)."""
    parts = [
        (reference or "").strip(),
        (nom or "").strip(),
        (prenom or "").strip(),
        fmt_v(telephone),
        fmt_v(village),
        fmt_v(prefecture),
        fmt_v(email),
        fmt_v(adresse),
        fmt_v(date_inscription),
    ]
    return " ".join(p for p in parts if p).strip()


def fmt_contribution_compact(montant: float, date_iso: str, note: str = "") -> str:
    """Représentation compacte d'une cotisation (sans labels)."""
    base = f"{float(montant):.2f} EUR {fmt_v(date_iso)}"
    n = (note or "").strip()
    return f"{base} {n}" if n else base


def fmt_depense_compact(description: str, montant: float, date_iso: str) -> str:
    """Représentation compacte d'une dépense (sans labels)."""
    return f"{fmt_v(description)} {float(montant):.2f} EUR {fmt_v(date_iso)}"


@st.cache_resource
def get_conn() -> PersistingConnection:
    db_file = bootstrap_local_database()
    raw = sqlite3.connect(db_file, check_same_thread=False)
    raw.row_factory = sqlite3.Row
    raw.execute("PRAGMA foreign_keys = ON;")
    conn = PersistingConnection(raw, db_file)
    init_db(conn)
    return conn


def reset_database_cache() -> None:
    clear_bootstrap_cache()
    get_conn.clear()


def render_storage_sidebar() -> None:
    with st.sidebar.expander("Sauvegarde des données", expanded=not persistence_is_configured()):
        cfg = get_persistence_config()
        pull_err = st.session_state.get("_persistence_pull_error")
        push_err = st.session_state.get("_persistence_push_error")

        if cfg:
            st.success("Stockage cloud actif — les données survivent aux redémarrages.")
            st.caption(f"Bucket : `{cfg['bucket']}` / `{cfg.get('key', 'agpm.db')}`")
        elif is_ephemeral_streamlit_host():
            st.error(
                "Sur Streamlit Cloud, le fichier `agpm.db` local est **effacé** à chaque redémarrage. "
                "Configurez un stockage cloud (voir ci-dessous) pour conserver vos données."
            )
        else:
            st.info(
                "En local, les données sont dans `agpm.db`. Sur Streamlit Cloud, ajoutez les secrets "
                "de persistance (fichier `.streamlit/secrets.toml.example` dans le dépôt)."
            )

        if pull_err:
            st.warning(f"Dernier chargement cloud : {pull_err}")
        if push_err:
            st.warning(f"Dernière sauvegarde cloud : {push_err}")

        db_path = local_db_path()
        if db_path.is_file():
            st.download_button(
                "Télécharger la base (agpm.db)",
                data=db_path.read_bytes(),
                file_name="agpm.db",
                mime="application/octet-stream",
                use_container_width=True,
            )
        else:
            st.caption("Aucun fichier de base local pour le moment.")

        uploaded = st.file_uploader("Restaurer une sauvegarde .db", type=["db"])
        if uploaded is not None and st.button("Appliquer la restauration", use_container_width=True):
            restore_database_file(uploaded.getvalue())
            reset_database_cache()
            st.success("Base restaurée. Rechargement…")
            st.rerun()

        if cfg and db_path.is_file() and st.button("Forcer l'envoi vers le cloud", use_container_width=True):
            try:
                push_db_to_cloud(cfg, db_path)
                st.session_state.pop("_persistence_push_error", None)
                st.success("Sauvegarde cloud mise à jour.")
            except Exception as exc:
                st.error(str(exc))

        with st.expander("Configurer Cloudflare R2 (gratuit)"):
            st.markdown(
                """
1. [Cloudflare](https://dash.cloudflare.com/) → **R2** → créer un bucket (ex. `agpm-gestion`).
2. **Manage R2 API Tokens** → token avec lecture/écriture sur ce bucket.
3. Sur [share.streamlit.io](https://share.streamlit.io/) → votre app → **Settings** → **Secrets** :
   collez le contenu de `.streamlit/secrets.toml.example` en remplissant les clés.
4. **Reboot** l'application.

Les modifications sont enregistrées automatiquement après chaque action dans l'app.
                """
            )


def init_db(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS membres (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            prenom TEXT NOT NULL,
            telephone TEXT NOT NULL DEFAULT '',
            village_origine TEXT NOT NULL DEFAULT '',
            adresse TEXT NOT NULL DEFAULT '',
            prefecture TEXT NOT NULL DEFAULT '',
            email TEXT NOT NULL DEFAULT '',
            date_inscription TEXT NOT NULL,
            actif INTEGER NOT NULL DEFAULT 1
        );
        """
    )
    # Homonymes autorises: on ne bloque plus nom/prenom/telephone.
    cur.execute("DROP INDEX IF EXISTS idx_member_identity;")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS contributions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            membre_id INTEGER NOT NULL,
            montant REAL NOT NULL CHECK (montant > 0),
            date TEXT NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            mode_paiement TEXT NOT NULL DEFAULT 'Espèces',
            FOREIGN KEY(membre_id) REFERENCES membres(id) ON DELETE CASCADE
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS depenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT NOT NULL,
            montant REAL NOT NULL CHECK (montant > 0),
            date TEXT NOT NULL
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS reports_membres (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            membre_id INTEGER NOT NULL,
            annee INTEGER NOT NULL,
            montant_du REAL NOT NULL DEFAULT 0 CHECK (montant_du >= 0),
            UNIQUE(membre_id, annee),
            FOREIGN KEY(membre_id) REFERENCES membres(id) ON DELETE CASCADE
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS reports_association (
            annee INTEGER PRIMARY KEY,
            solde_reporte REAL NOT NULL DEFAULT 0
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS activites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            type_operation TEXT NOT NULL,
            entite TEXT NOT NULL,
            entite_id INTEGER,
            details TEXT NOT NULL
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS parametres (
            cle TEXT PRIMARY KEY,
            valeur TEXT NOT NULL
        );
        """
    )
    # Migration legere: ajoute un identifiant membre unique et stable.
    cols = [row["name"] for row in conn.execute("PRAGMA table_info(membres)").fetchall()]
    if "reference" not in cols:
        conn.execute("ALTER TABLE membres ADD COLUMN reference TEXT;")
    if "village_origine" not in cols:
        conn.execute("ALTER TABLE membres ADD COLUMN village_origine TEXT NOT NULL DEFAULT '';")
    contrib_cols = [row["name"] for row in conn.execute("PRAGMA table_info(contributions)").fetchall()]
    if "mode_paiement" not in contrib_cols:
        conn.execute(
            "ALTER TABLE contributions ADD COLUMN mode_paiement TEXT NOT NULL DEFAULT 'Espèces';"
        )
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_membres_reference ON membres(reference);")
    # Normalise toutes les references au format M007.
    conn.execute("UPDATE membres SET reference = printf('M%03d', id);")
    conn.commit()
    # Migration report dynamique : purge unique des reports postérieurs à l'année de
    # référence (devenus obsolètes, désormais recalculés depuis les cotisations).
    already = conn.execute(
        "SELECT valeur FROM parametres WHERE cle = 'reports_purged'"
    ).fetchone()
    if not already:
        base = get_baseline_year(conn)
        removed = conn.execute(
            "DELETE FROM reports_membres WHERE annee > ?;", (base,)
        ).rowcount
        conn.execute(
            """
            INSERT INTO parametres(cle, valeur) VALUES('reports_purged', '1')
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur;
            """
        )
        if removed:
            log_activity(
                conn,
                type_operation="DELETE",
                entite="reports_membres",
                entite_id=None,
                details=(
                    f"Purge migration report dynamique : {removed} report(s) postérieur(s) "
                    f"à {base} supprimé(s)."
                ),
            )
        conn.commit()


def to_iso(value: date | datetime) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value.isoformat()


def fetch_df(conn: sqlite3.Connection, query: str, params: tuple = ()) -> pd.DataFrame:
    return pd.read_sql_query(query, conn, params=params)


def month_diff_inclusive(start_date: date, end_date: date) -> int:
    if end_date < start_date:
        return 0
    return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month) + 1


def expected_months_for_member(inscription: date, year: int) -> int:
    start = max(inscription, date(year, 1, 1))
    end = date(year, 12, 31)
    today = date.today()
    if year == today.year:
        end = min(end, today)
    if year > today.year:
        return 0
    return month_diff_inclusive(start, end)


def first_sunday(year: int, month: int) -> date:
    """Date de réunion = 1er dimanche du mois."""
    d = date(year, month, 1)
    # weekday(): lundi=0 … dimanche=6
    return d + timedelta(days=(6 - d.weekday()) % 7)


def ensure_parametres_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS parametres (
            cle TEXT PRIMARY KEY,
            valeur TEXT NOT NULL
        );
        """
    )


def get_monthly_contribution(conn: sqlite3.Connection) -> float:
    try:
        row = conn.execute(
            "SELECT valeur FROM parametres WHERE cle = 'montant_cotisation'"
        ).fetchone()
    except sqlite3.OperationalError:
        ensure_parametres_table(conn)
        conn.commit()
        return DEFAULT_MONTHLY_CONTRIBUTION
    if row and row["valeur"]:
        try:
            return float(row["valeur"])
        except (TypeError, ValueError):
            pass
    return DEFAULT_MONTHLY_CONTRIBUTION


def set_monthly_contribution(conn: sqlite3.Connection, amount: float) -> None:
    ensure_parametres_table(conn)
    conn.execute(
        """
        INSERT INTO parametres(cle, valeur)
        VALUES('montant_cotisation', ?)
        ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur;
        """,
        (str(float(amount)),),
    )
    log_activity(
        conn,
        type_operation="UPDATE",
        entite="parametres",
        entite_id=None,
        details=f"Montant cotisation mensuelle defini a {float(amount):.2f} EUR",
    )
    conn.commit()


def get_association_report(conn: sqlite3.Connection, year: int) -> float:
    row = conn.execute(
        "SELECT solde_reporte FROM reports_association WHERE annee = ?",
        (year,),
    ).fetchone()
    return float(row["solde_reporte"]) if row else 0.0


def upsert_association_report(conn: sqlite3.Connection, year: int, amount: float) -> None:
    conn.execute(
        """
        INSERT INTO reports_association(annee, solde_reporte)
        VALUES(?, ?)
        ON CONFLICT(annee) DO UPDATE SET solde_reporte = excluded.solde_reporte;
        """,
        (year, amount),
    )
    log_activity(
        conn,
        type_operation="UPDATE",
        entite="reports_association",
        entite_id=year,
        details=f"Solde report association {year} {amount:.2f} EUR",
    )
    conn.commit()


def upsert_member_report(conn: sqlite3.Connection, member_id: int, year: int, amount_due: float) -> None:
    conn.execute(
        """
        INSERT INTO reports_membres(membre_id, annee, montant_du)
        VALUES(?, ?, ?)
        ON CONFLICT(membre_id, annee) DO UPDATE SET montant_du = excluded.montant_du;
        """,
        (member_id, year, amount_due),
    )
    log_activity(
        conn,
        type_operation="UPDATE",
        entite="reports_membres",
        entite_id=member_id,
        details=f"Solde N-1 membre {member_ref(member_id)} exercice {year} {amount_due:.2f} EUR",
    )
    conn.commit()


def log_activity(
    conn: sqlite3.Connection,
    type_operation: str,
    entite: str,
    entite_id: Optional[int],
    details: str,
) -> None:
    conn.execute(
        """
        INSERT INTO activites(created_at, type_operation, entite, entite_id, details)
        VALUES(?, ?, ?, ?, ?)
        """,
        (datetime.now().isoformat(timespec="seconds"), type_operation, entite, entite_id, details),
    )


def total_contributions(conn: sqlite3.Connection, year: Optional[int] = None) -> float:
    if year is None:
        row = conn.execute("SELECT COALESCE(SUM(montant), 0) AS total FROM contributions").fetchone()
    else:
        row = conn.execute(
            """
            SELECT COALESCE(SUM(montant), 0) AS total
            FROM contributions
            WHERE strftime('%Y', date) = ?
            """,
            (str(year),),
        ).fetchone()
    return float(row["total"])


def total_expenses(conn: sqlite3.Connection, year: Optional[int] = None) -> float:
    if year is None:
        row = conn.execute("SELECT COALESCE(SUM(montant), 0) AS total FROM depenses").fetchone()
    else:
        row = conn.execute(
            """
            SELECT COALESCE(SUM(montant), 0) AS total
            FROM depenses
            WHERE strftime('%Y', date) = ?
            """,
            (str(year),),
        ).fetchone()
    return float(row["total"])


def get_baseline_year(conn: sqlite3.Connection) -> int:
    """Premier exercice géré : ancre la dette initiale (Solde N-1 importé)."""
    row = conn.execute("SELECT MIN(annee) AS y FROM reports_membres").fetchone()
    if row and row["y"] is not None:
        return int(row["y"])
    row = conn.execute(
        "SELECT MIN(CAST(strftime('%Y', date) AS INTEGER)) AS y FROM contributions"
    ).fetchone()
    if row and row["y"] is not None:
        return int(row["y"])
    return date.today().year


def full_year_months(inscription: date, year: int) -> int:
    """Mois dus sur l'année complète (Jan→Déc), pondérés par la date d'inscription."""
    start = max(inscription, date(year, 1, 1))
    end = date(year, 12, 31)
    if start > end:
        return 0
    return month_diff_inclusive(start, end)


def get_members_status(conn: sqlite3.Connection, year: int, include_archived: bool = False) -> pd.DataFrame:
    where_clause = "" if include_archived else "WHERE actif = 1"
    members = fetch_df(
        conn,
        """
        SELECT id, reference, nom, prenom, telephone, village_origine, email, date_inscription, actif
        FROM membres
        """
        + where_clause
        + """
        ORDER BY nom, prenom;
        """,
    )
    cols_out = [
        "id",
        "reference",
        "actif",
        "nom",
        "prenom",
        "telephone",
        "village_origine",
        "email",
        "solde_n1",
        "total_paye",
        "attendu",
        "reste",
        "statut",
    ]
    if members.empty:
        return pd.DataFrame(columns=cols_out)

    monthly_amount = get_monthly_contribution(conn)
    baseline_year = get_baseline_year(conn)

    # Dette initiale (arriérés d'avant le premier exercice géré), par membre.
    baseline = fetch_df(
        conn,
        "SELECT membre_id, montant_du FROM reports_membres WHERE annee = ?;",
        (baseline_year,),
    )
    baseline_map = {int(r["membre_id"]): float(r["montant_du"]) for _, r in baseline.iterrows()}

    # Paiements par membre et par année : servent à dériver les reports.
    paid_year = fetch_df(
        conn,
        """
        SELECT membre_id,
               CAST(strftime('%Y', date) AS INTEGER) AS annee,
               COALESCE(SUM(montant), 0) AS paye
        FROM contributions
        GROUP BY membre_id, annee;
        """,
    )
    paid_map = {
        (int(r["membre_id"]), int(r["annee"])): float(r["paye"])
        for _, r in paid_year.iterrows()
    }

    members["date_inscription"] = pd.to_datetime(members["date_inscription"], errors="coerce").dt.date

    rows = []
    for _, mem in members.iterrows():
        mid = int(mem["id"])
        inscription = mem["date_inscription"] if pd.notna(mem["date_inscription"]) else date(year, 1, 1)

        # Report entrant = dette initiale + reliquats des exercices précédents.
        # Pas de plancher : un trop-perçu se reporte en crédit (valeur négative).
        if year < baseline_year:
            carry_in = get_member_report(conn, mid, year)
        else:
            carry_in = baseline_map.get(mid, 0.0)
            for k in range(baseline_year, year):
                carry_in += full_year_months(inscription, k) * monthly_amount
                carry_in -= paid_map.get((mid, k), 0.0)

        total_paye = paid_map.get((mid, year), 0.0)
        attendu = carry_in + expected_months_for_member(inscription, year) * monthly_amount
        reste = attendu - total_paye
        statut = "A jour" if reste <= 0.001 else "En retard"

        rows.append(
            {
                "id": mid,
                "reference": mem["reference"],
                "actif": mem["actif"],
                "nom": mem["nom"],
                "prenom": mem["prenom"],
                "telephone": mem["telephone"],
                "village_origine": mem["village_origine"] if "village_origine" in mem else "",
                "email": mem["email"],
                "solde_n1": round(carry_in, 2),
                "total_paye": round(total_paye, 2),
                "attendu": round(attendu, 2),
                "reste": round(reste, 2),
                "statut": statut,
            }
        )

    return pd.DataFrame(rows, columns=cols_out)


def get_member_report(conn: sqlite3.Connection, member_id: int, year: int) -> float:
    row = conn.execute(
        "SELECT montant_du FROM reports_membres WHERE membre_id = ? AND annee = ?",
        (member_id, year),
    ).fetchone()
    return float(row["montant_du"]) if row else 0.0


def status_display_df(status: pd.DataFrame) -> pd.DataFrame:
    cols = {
        "reference": "Référence",
        "nom": "Nom",
        "prenom": "Prénom",
        "telephone": "Téléphone",
        "village_origine": "Village",
        "email": "Email",
        "solde_n1": "Solde N-1",
        "total_paye": "Payé",
        "attendu": "Attendu",
        "reste": "Reste",
        "statut": "Statut",
    }
    keep = [c for c in cols if c in status.columns]
    return status[keep].rename(columns=cols)


def format_eur(amount: float) -> str:
    return f"{amount:,.2f} EUR".replace(",", " ")


def member_pick_label(row: pd.Series) -> str:
    reste = float(row.get("reste", 0))
    statut = str(row.get("statut", ""))
    icon = "⏳" if statut == "En retard" else "✅"
    return (
        f"{icon} {row['reference']} — {row['nom']} {row['prenom']} "
        f"(reste {format_eur(reste)})"
    )


def filter_members_status(
    status: pd.DataFrame,
    search: str,
    only_late: bool,
) -> pd.DataFrame:
    if status.empty:
        return status
    out = status.copy()
    if only_late:
        out = out[out["statut"] == "En retard"]
    q = search.strip().lower()
    if q:
        mask = (
            out["reference"].astype(str).str.lower().str.contains(q, na=False)
            | out["nom"].astype(str).str.lower().str.contains(q, na=False)
            | out["prenom"].astype(str).str.lower().str.contains(q, na=False)
            | out["telephone"].astype(str).str.lower().str.contains(q, na=False)
            | out["village_origine"].astype(str).str.lower().str.contains(q, na=False)
        )
        out = out[mask]
    return out


def sort_status_for_entry(status: pd.DataFrame) -> pd.DataFrame:
    if status.empty:
        return status
    late_first = (status["statut"] != "En retard").astype(int)
    return (
        status.assign(_late=late_first)
        .sort_values(["_late", "reste", "nom", "prenom"], ascending=[True, False, True, True])
        .drop(columns="_late")
        .reset_index(drop=True)
    )


def status_grid_dataframe(status: pd.DataFrame) -> pd.DataFrame:
    """Tableau affiché : lisible, avec repères visuels pour le statut."""
    if status.empty:
        return pd.DataFrame()
    view = status.copy()
    view["statut"] = view["statut"].map({"A jour": "✅ À jour", "En retard": "⏳ En retard"}).fillna(view["statut"])
    for col in ("solde_n1", "total_paye", "attendu", "reste"):
        if col in view.columns:
            view[col] = view[col].map(lambda x: round(float(x), 2))
    return status_display_df(view)


def insert_contribution(
    conn: sqlite3.Connection,
    member_id: int,
    montant: float,
    contribution_date: date,
    note: str = "",
    mode_paiement: str = DEFAULT_PAYMENT_METHOD,
) -> int:
    mode = mode_paiement if mode_paiement in PAYMENT_METHODS else DEFAULT_PAYMENT_METHOD
    cur = conn.execute(
        "INSERT INTO contributions(membre_id, montant, date, note, mode_paiement) VALUES(?, ?, ?, ?, ?)",
        (member_id, float(montant), to_iso(contribution_date), note.strip(), mode),
    )
    cid = int(cur.lastrowid)
    log_activity(
        conn,
        type_operation="CREATE",
        entite="contribution",
        entite_id=cid,
        details=(
            f"Ajout cotisation #{cid} {member_ref(member_id)} "
            + fmt_contribution_compact(float(montant), to_iso(contribution_date), note)
            + f" [{mode}]"
        ),
    )
    conn.commit()
    return cid


def render_app_styles() -> None:
    st.markdown(
        """
        <style>
        .page-guide {
            font-size: 0.95rem;
            color: #475569;
            margin: 0 0 0.75rem 0;
            line-height: 1.45;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            padding: 0.35rem 0.75rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_page_guide(text: str) -> None:
    st.markdown(f'<p class="page-guide">{text}</p>', unsafe_allow_html=True)


def filter_df_search(df: pd.DataFrame, search: str, columns: list[str]) -> pd.DataFrame:
    q = search.strip().lower()
    if not q or df.empty:
        return df
    mask = pd.Series(False, index=df.index)
    for col in columns:
        if col in df.columns:
            mask = mask | df[col].astype(str).str.lower().str.contains(q, na=False)
    return df[mask]


def member_row_label(reference: str, nom: str, prenom: str, extra: str = "") -> str:
    base = f"{reference} — {nom} {prenom}"
    return f"{base} · {extra}" if extra else base


def insert_depense(
    conn: sqlite3.Connection,
    description: str,
    montant: float,
    depense_date: date,
) -> int:
    conn.execute(
        "INSERT INTO depenses(description, montant, date) VALUES(?, ?, ?)",
        (description.strip(), float(montant), to_iso(depense_date)),
    )
    dep_id = int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
    log_activity(
        conn,
        type_operation="CREATE",
        entite="depense",
        entite_id=dep_id,
        details=f"Ajout depense #{dep_id} " + fmt_depense_compact(description, float(montant), to_iso(depense_date)),
    )
    conn.commit()
    return dep_id


def activity_entite_label(entite: str) -> str:
    labels = {
        "membre": "Membre",
        "contribution": "Cotisation",
        "depense": "Dépense",
        "reports_membres": "Solde N-1 membre",
        "reports_association": "Solde association",
        "parametres": "Paramètre",
        "import_excel": "Import Excel",
    }
    return labels.get(str(entite), str(entite))


def get_default_country_code(conn: sqlite3.Connection) -> str:
    try:
        row = conn.execute(
            "SELECT valeur FROM parametres WHERE cle = 'indicatif_tel'"
        ).fetchone()
    except sqlite3.OperationalError:
        ensure_parametres_table(conn)
        conn.commit()
        return "33"
    if row and row["valeur"]:
        digits = re.sub(r"\D", "", str(row["valeur"]))
        if digits:
            return digits
    return "33"


def set_default_country_code(conn: sqlite3.Connection, code: str) -> None:
    ensure_parametres_table(conn)
    digits = re.sub(r"\D", "", str(code)) or "33"
    conn.execute(
        """
        INSERT INTO parametres(cle, valeur)
        VALUES('indicatif_tel', ?)
        ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur;
        """,
        (digits,),
    )
    conn.commit()


def phone_to_intl(phone: object, country_code: str) -> str:
    """Convertit un numéro local en format international (chiffres uniquement)."""
    digits = re.sub(r"\D", "", str(phone or ""))
    if not digits:
        return ""
    cc = re.sub(r"\D", "", country_code) or "33"
    if digits.startswith("00"):
        return digits[2:]
    if digits.startswith(cc):
        return digits
    digits = digits.lstrip("0")
    return cc + digits


def whatsapp_link(phone: object, message: str, country_code: str) -> str:
    intl = phone_to_intl(phone, country_code)
    if not intl:
        return ""
    return f"https://wa.me/{intl}?text={urllib.parse.quote(message)}"


def mailto_link(email: object, subject: str, body: str) -> str:
    addr = str(email or "").strip()
    if not addr or "@" not in addr:
        return ""
    query = urllib.parse.urlencode({"subject": subject, "body": body})
    return f"mailto:{addr}?{query}"


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Export") -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    return buf.getvalue()


def render_export_buttons(df: pd.DataFrame, basename: str, key: str) -> None:
    if df.empty:
        return
    today = date.today().isoformat()
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "⬇️ CSV",
            data=df_to_csv_bytes(df),
            file_name=f"{basename}_{today}.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"{key}_csv",
        )
    with c2:
        st.download_button(
            "⬇️ Excel",
            data=df_to_excel_bytes(df),
            file_name=f"{basename}_{today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{key}_xlsx",
        )


def build_receipt_pdf(
    *,
    reference: str,
    nom: str,
    prenom: str,
    montant: float,
    date_paiement: str,
    note: str,
    annee: int,
    total_paye_annee: float,
    reste_annee: float,
    mode_paiement: str = DEFAULT_PAYMENT_METHOD,
) -> Optional[bytes]:
    """Génère un reçu PDF. Renvoie None si fpdf2 n'est pas installé."""
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    pdf = FPDF(format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    if LOGO_PATH.is_file():
        try:
            pdf.image(str(LOGO_PATH), x=15, y=12, w=38)
        except Exception:
            pass

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_xy(60, 16)
    pdf.cell(0, 8, "AGPM", ln=1)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_x(60)
    pdf.cell(0, 6, "Association des Guineens du Pays de Meaux", ln=1)

    pdf.ln(18)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "RECU DE COTISATION", ln=1, align="C")
    pdf.set_draw_color(37, 99, 235)
    pdf.set_line_width(0.6)
    y = pdf.get_y()
    pdf.line(15, y, 195, y)
    pdf.ln(8)

    def _safe(txt: str) -> str:
        return str(txt).encode("latin-1", "replace").decode("latin-1")

    rows = [
        ("Membre", f"{reference} - {nom} {prenom}"),
        ("Date du paiement", date_paiement),
        ("Montant recu", f"{montant:.2f} EUR"),
        ("Mode de paiement", mode_paiement or DEFAULT_PAYMENT_METHOD),
        ("Objet", note or "Cotisation"),
        ("Exercice", str(annee)),
        ("Total paye (annee)", f"{total_paye_annee:.2f} EUR"),
        ("Reste du (annee)", f"{reste_annee:.2f} EUR"),
    ]
    pdf.set_font("Helvetica", "", 11)
    for label, value in rows:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(55, 9, _safe(label), border=0)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 9, _safe(value), ln=1)

    pdf.ln(12)
    pdf.set_font("Helvetica", "I", 9)
    pdf.multi_cell(
        0,
        5,
        _safe(
            f"Recu emis le {date.today().strftime('%d/%m/%Y')} par l'application de gestion AGPM. "
            "Document a conserver comme preuve de paiement."
        ),
    )

    out = pdf.output()
    return bytes(out)


def _render_member_edit_panel(conn: sqlite3.Connection, selected_id: int) -> None:
    member_row = conn.execute(
        """
        SELECT id, actif, reference, nom, prenom, telephone, village_origine,
               adresse, prefecture, email, date_inscription
        FROM membres WHERE id = ?
        """,
        (selected_id,),
    ).fetchone()
    if not member_row:
        return

    etat = "Actif" if int(member_row["actif"]) == 1 else "Archivé"
    with st.container(border=True):
        st.markdown(f"#### {member_row['reference']} — {member_row['nom']} {member_row['prenom']}")
        st.caption(f"État : **{etat}**")

        try:
            current_inscription = datetime.fromisoformat(member_row["date_inscription"]).date()
        except (TypeError, ValueError):
            current_inscription = date.today()

        with st.form(f"edit_member_{selected_id}"):
            e1, e2 = st.columns(2)
            with e1:
                edit_nom = st.text_input("Nom *", value=member_row["nom"] or "")
                edit_prenom = st.text_input("Prénom *", value=member_row["prenom"] or "")
                edit_telephone = st.text_input("Téléphone", value=member_row["telephone"] or "")
                edit_village = st.text_input("Village d'origine", value=member_row["village_origine"] or "")
                edit_prefecture = st.text_input("Préfecture", value=member_row["prefecture"] or "")
            with e2:
                edit_email = st.text_input("Email", value=member_row["email"] or "")
                edit_adresse = st.text_input("Adresse", value=member_row["adresse"] or "")
                edit_inscription = st.date_input("Date d'inscription", value=current_inscription)

            if st.form_submit_button("Enregistrer les modifications", use_container_width=True):
                nom_v = edit_nom.strip()
                prenom_v = edit_prenom.strip()
                if not nom_v or not prenom_v:
                    st.error("Nom et prénom sont obligatoires.")
                elif edit_email.strip() and (
                    "@" not in edit_email or "." not in edit_email.strip().split("@")[-1]
                ):
                    st.error("Email invalide.")
                else:
                    before = fmt_member_compact(
                        reference=member_row["reference"],
                        nom=member_row["nom"],
                        prenom=member_row["prenom"],
                        telephone=member_row["telephone"],
                        village=member_row["village_origine"],
                        prefecture=member_row["prefecture"],
                        email=member_row["email"],
                        adresse=member_row["adresse"],
                        date_inscription=member_row["date_inscription"],
                    )
                    conn.execute(
                        """
                        UPDATE membres
                        SET nom = ?, prenom = ?, telephone = ?, village_origine = ?,
                            adresse = ?, prefecture = ?, email = ?, date_inscription = ?
                        WHERE id = ?
                        """,
                        (
                            nom_v,
                            prenom_v.strip(),
                            edit_telephone.strip(),
                            edit_village.strip(),
                            edit_adresse.strip(),
                            edit_prefecture.strip(),
                            edit_email.strip(),
                            to_iso(edit_inscription),
                            selected_id,
                        ),
                    )
                    after = fmt_member_compact(
                        reference=member_row["reference"],
                        nom=nom_v,
                        prenom=prenom_v,
                        telephone=edit_telephone.strip(),
                        village=edit_village.strip(),
                        prefecture=edit_prefecture.strip(),
                        email=edit_email.strip(),
                        adresse=edit_adresse.strip(),
                        date_inscription=to_iso(edit_inscription),
                    )
                    log_activity(
                        conn,
                        type_operation="UPDATE",
                        entite="membre",
                        entite_id=selected_id,
                        details=f"Maj membre {before}  ->  {after}",
                    )
                    conn.commit()
                    st.success("Fiche membre mise à jour.")
                    st.rerun()

        with st.expander("Dette initiale (Solde N-1)", expanded=False):
            report_year = get_baseline_year(conn)
            st.caption(
                f"Arriérés repris **avant** le premier exercice géré ({report_year}). "
                "Les exercices suivants sont recalculés automatiquement à partir des cotisations."
            )
            current_solde = get_member_report(conn, selected_id, int(report_year))
            new_solde = st.number_input(
                f"Montant dû avant {report_year} (EUR)",
                min_value=0.0,
                value=float(current_solde),
                step=1.0,
                format="%.2f",
                key=f"member_solde_n1_{selected_id}_{int(report_year)}",
            )
            if st.button("Enregistrer le solde N-1", key=f"save_solde_n1_{selected_id}"):
                upsert_member_report(conn, selected_id, int(report_year), float(new_solde))
                st.toast("Solde N-1 enregistré.")
                st.rerun()

        a1, a2 = st.columns(2)
        with a1:
            if int(member_row["actif"]) == 1:
                if st.button("Archiver ce membre", type="secondary", use_container_width=True):
                    conn.execute("UPDATE membres SET actif = 0 WHERE id = ?", (selected_id,))
                    log_activity(
                        conn,
                        type_operation="UPDATE",
                        entite="membre",
                        entite_id=selected_id,
                        details=f"Archivage membre {member_row['reference']} {member_row['nom']} {member_row['prenom']}",
                    )
                    conn.commit()
                    st.rerun()
        with a2:
            if int(member_row["actif"]) == 0:
                if st.button("Réactiver ce membre", use_container_width=True):
                    conn.execute("UPDATE membres SET actif = 1 WHERE id = ?", (selected_id,))
                    log_activity(
                        conn,
                        type_operation="UPDATE",
                        entite="membre",
                        entite_id=selected_id,
                        details=f"Reactivation membre {member_row['reference']} {member_row['nom']} {member_row['prenom']}",
                    )
                    conn.commit()
                    st.rerun()


def page_membres(conn: sqlite3.Connection) -> None:
    st.subheader("Membres")
    render_page_guide(
        "1️⃣ Cliquez sur un membre dans le tableau &nbsp;→&nbsp; 2️⃣ Modifiez sa fiche à droite. "
        "Onglet <strong>Soldes N-1</strong> pour l'initialisation en masse."
    )

    counts = conn.execute(
        """
        SELECT
            SUM(CASE WHEN actif = 1 THEN 1 ELSE 0 END) AS actifs,
            SUM(CASE WHEN actif = 0 THEN 1 ELSE 0 END) AS archives
        FROM membres;
        """
    ).fetchone()
    k1, k2, k3 = st.columns(3)
    k1.metric("Membres actifs", int(counts["actifs"] or 0))
    k2.metric("Archivés", int(counts["archives"] or 0))
    k3.metric("Total", int((counts["actifs"] or 0) + (counts["archives"] or 0)))

    tab_liste, tab_ajouter, tab_soldes = st.tabs(["📋 Liste", "➕ Nouveau", "💰 Soldes N-1"])

    with tab_ajouter:
        with st.container(border=True):
            st.markdown("##### Ajouter un membre")
            with st.form("add_member", clear_on_submit=True):
                c1, c2 = st.columns(2)
                with c1:
                    nom = st.text_input("Nom *")
                    prenom = st.text_input("Prénom *")
                    telephone = st.text_input("Téléphone")
                    village_origine = st.text_input("Village d'origine")
                    prefecture = st.text_input("Préfecture")
                with c2:
                    email = st.text_input("Email")
                    adresse = st.text_input("Adresse")
                    date_inscription = st.date_input("Date d'inscription", value=date.today())
                if st.form_submit_button("Ajouter le membre", type="primary", use_container_width=True):
                    nom_v, prenom_v = nom.strip(), prenom.strip()
                    email_v = email.strip()
                    if not nom_v or not prenom_v:
                        st.error("Nom et prénom sont obligatoires.")
                    elif email_v and ("@" not in email_v or "." not in email_v.split("@")[-1]):
                        st.error("Email invalide.")
                    else:
                        try:
                            cur = conn.execute(
                                """
                                INSERT INTO membres(reference, nom, prenom, telephone, village_origine,
                                    adresse, prefecture, email, date_inscription, actif)
                                VALUES('', ?, ?, ?, ?, ?, ?, ?, ?, 1);
                                """,
                                (
                                    nom_v,
                                    prenom_v,
                                    telephone.strip(),
                                    village_origine.strip(),
                                    adresse.strip(),
                                    prefecture.strip(),
                                    email_v,
                                    to_iso(date_inscription),
                                ),
                            )
                            new_id = int(cur.lastrowid)
                            conn.execute(
                                "UPDATE membres SET reference = ? WHERE id = ?",
                                (member_ref(new_id), new_id),
                            )
                            log_activity(
                                conn,
                                type_operation="CREATE",
                                entite="membre",
                                entite_id=new_id,
                                details="Ajout membre "
                                + fmt_member_compact(
                                    reference=member_ref(new_id),
                                    nom=nom_v,
                                    prenom=prenom_v,
                                    telephone=telephone.strip(),
                                    village=village_origine.strip(),
                                    prefecture=prefecture.strip(),
                                    email=email_v,
                                    adresse=adresse.strip(),
                                    date_inscription=to_iso(date_inscription),
                                ),
                            )
                            conn.commit()
                            st.session_state["membre_selected_id"] = new_id
                            st.success(f"Membre {member_ref(new_id)} ajouté.")
                            st.rerun()
                        except sqlite3.IntegrityError:
                            st.error("Impossible d'ajouter ce membre.")

    with tab_soldes:
        init_year = get_baseline_year(conn)
        st.caption(
            f"Dette initiale par membre, reprise **avant** le premier exercice géré ({init_year}) "
            "— colonne « Solde N-1 » de l'Excel. Les reports des années suivantes sont "
            "recalculés automatiquement à partir des cotisations enregistrées."
        )
        init_members = fetch_df(
            conn,
            """
            SELECT id, reference, nom, prenom
            FROM membres WHERE actif = 1 ORDER BY nom, prenom;
            """,
        )
        if init_members.empty:
            st.info("Aucun membre actif.")
        else:
            init_rows = [
                {
                    "id": int(r["id"]),
                    "reference": r["reference"],
                    "nom": r["nom"],
                    "prenom": r["prenom"],
                    "solde_n1": get_member_report(conn, int(r["id"]), int(init_year)),
                }
                for _, r in init_members.iterrows()
            ]
            edited_init = st.data_editor(
                pd.DataFrame(init_rows),
                column_config={
                    "id": st.column_config.NumberColumn("Id", disabled=True),
                    "reference": st.column_config.TextColumn("Référence", disabled=True),
                    "nom": st.column_config.TextColumn("Nom", disabled=True),
                    "prenom": st.column_config.TextColumn("Prénom", disabled=True),
                    "solde_n1": st.column_config.NumberColumn(
                        f"Solde N-1 ({int(init_year)})",
                        min_value=0.0,
                        step=1.0,
                        format="%.2f",
                    ),
                },
                hide_index=True,
                num_rows="fixed",
                use_container_width=True,
                key=f"bulk_solde_editor_{int(init_year)}",
            )
            if st.button("Enregistrer tous les soldes N-1", type="primary", key="bulk_save_solde_n1"):
                for _, row in edited_init.iterrows():
                    upsert_member_report(conn, int(row["id"]), int(init_year), float(row["solde_n1"]))
                st.success(f"{len(edited_init)} solde(s) enregistré(s).")
                st.rerun()

    with tab_liste:
        view_mode = st.radio(
            "Afficher",
            ["Actifs", "Archivés", "Tous"],
            horizontal=True,
            index=0,
            key="membres_view_mode",
        )
        where_sql = "WHERE actif = 1"
        if view_mode == "Archivés":
            where_sql = "WHERE actif = 0"
        elif view_mode == "Tous":
            where_sql = ""

        df = fetch_df(
            conn,
            """
            SELECT id, reference,
                   CASE WHEN actif = 1 THEN '✅ Actif' ELSE '📦 Archivé' END AS etat,
                   nom, prenom, telephone, village_origine, prefecture, email, date_inscription
            FROM membres
            """
            + where_sql
            + " ORDER BY nom, prenom;",
        )
        if df.empty:
            st.info("Aucun membre dans cette catégorie.")
            return

        search = st.text_input(
            "Rechercher",
            "",
            key="membres_search",
            placeholder="Nom, référence, téléphone, village…",
        )
        df = filter_df_search(
            df,
            search,
            ["reference", "nom", "prenom", "telephone", "village_origine", "prefecture"],
        )
        st.caption(f"{len(df)} membre(s) — cliquez sur une ligne pour ouvrir la fiche.")

        grid_view = df.drop(columns=["id"], errors="ignore")
        id_by_row = df["id"].astype(int).tolist()

        selection = st.dataframe(
            grid_view,
            use_container_width=True,
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row",
            key="membres_grid",
            column_config={
                "reference": st.column_config.TextColumn("Réf.", width="small"),
                "etat": st.column_config.TextColumn("État", width="small"),
            },
        )

        selected_rows = []
        if selection is not None and hasattr(selection, "selection") and selection.selection is not None:
            selected_rows = list(selection.selection.rows or [])

        if selected_rows and 0 <= selected_rows[0] < len(id_by_row):
            st.session_state["membre_selected_id"] = id_by_row[selected_rows[0]]

        valid_ids = set(id_by_row)
        if "membre_selected_id" not in st.session_state or st.session_state["membre_selected_id"] not in valid_ids:
            st.session_state["membre_selected_id"] = id_by_row[0]

        pick_options = {
            member_row_label(
                str(r["reference"]),
                str(r["nom"]),
                str(r["prenom"]),
                str(r["village_origine"] or r["telephone"] or ""),
            ): int(r["id"])
            for _, r in df.iterrows()
        }
        pick_labels = list(pick_options.keys())
        current_id = int(st.session_state["membre_selected_id"])
        pick_idx = next((i for i, lbl in enumerate(pick_labels) if pick_options[lbl] == current_id), 0)
        picked = st.selectbox("Membre sélectionné", pick_labels, index=pick_idx)
        st.session_state["membre_selected_id"] = pick_options[picked]

        st.divider()
        _render_member_edit_panel(conn, int(st.session_state["membre_selected_id"]))


def page_contributions(conn: sqlite3.Connection) -> None:
    st.subheader("Cotisations")
    render_page_guide("1️⃣ Choisissez un membre dans le tableau → 2️⃣ Enregistrez sa cotisation à droite.")

    members = fetch_df(
        conn,
        "SELECT id FROM membres WHERE actif = 1",
    )
    if members.empty:
        st.info("Commencez par ajouter au moins un membre (menu **Membres**).")
        return

    year = st.number_input(
        "Exercice",
        min_value=2020,
        max_value=2100,
        value=date.today().year,
        step=1,
        key="contrib_year",
        help="Année prise en compte pour le statut et l'historique.",
    )
    year_int = int(year)
    monthly_amount = get_monthly_contribution(conn)

    status_all = get_members_status(conn, year_int)
    if status_all.empty:
        st.info("Aucun membre actif.")
        return

    nb_late = int((status_all["statut"] == "En retard").sum())
    nb_ok = int(len(status_all) - nb_late)
    total_year = total_contributions(conn, year_int)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Membres actifs", len(status_all))
    k2.metric("À jour", nb_ok)
    k3.metric("En retard", nb_late)
    k4.metric("Encaissé (année)", format_eur(total_year))

    if nb_late > 0:
        with st.expander(f"📣 Relancer les retardataires ({nb_late})", expanded=False):
            country_code = get_default_country_code(conn)
            late_df = status_all[status_all["statut"] == "En retard"].copy()
            default_msg = (
                "Bonjour {prenom}, petit rappel amical : votre cotisation AGPM {annee} "
                "presente un reste de {reste} EUR. Merci de regulariser lors de la prochaine reunion. "
                "Cordialement, le bureau AGPM."
            )
            msg_template = st.text_area(
                "Message de rappel",
                value=default_msg,
                help="Variables disponibles : {prenom}, {nom}, {reste}, {annee}.",
                key="reminder_template",
            )

            def _msg_for(r: pd.Series) -> str:
                try:
                    return msg_template.format(
                        prenom=r["prenom"],
                        nom=r["nom"],
                        reste=f"{float(r['reste']):.2f}",
                        annee=year_int,
                    )
                except (KeyError, ValueError):
                    return msg_template

            rows = []
            for _, r in late_df.iterrows():
                message = _msg_for(r)
                subject = f"Rappel cotisation AGPM {year_int}"
                rows.append(
                    {
                        "Référence": r["reference"],
                        "Nom": f"{r['nom']} {r['prenom']}",
                        "Téléphone": r["telephone"] or "—",
                        "Reste": float(r["reste"]),
                        "WhatsApp": whatsapp_link(r["telephone"], message, country_code),
                        "Email": mailto_link(r["email"], subject, message),
                    }
                )
            reminders = pd.DataFrame(rows)

            st.caption(
                f"Indicatif téléphonique : +{country_code} "
                "(modifiable dans Tableau de bord → Paramètres). "
                "Cliquez sur un lien pour ouvrir WhatsApp ou l'email pré-rempli."
            )
            st.dataframe(
                reminders,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Reste": st.column_config.NumberColumn(format="%.2f EUR"),
                    "WhatsApp": st.column_config.LinkColumn("WhatsApp", display_text="Envoyer"),
                    "Email": st.column_config.LinkColumn("Email", display_text="Écrire"),
                },
            )
            render_export_buttons(
                reminders.drop(columns=["WhatsApp", "Email"]),
                f"retardataires_{year_int}",
                key="export_reminders",
            )

    f1, f2 = st.columns([3, 1])
    with f1:
        search = st.text_input(
            "Rechercher un membre",
            "",
            key="contrib_search",
            placeholder="Nom, référence M005, téléphone, village…",
        )
    with f2:
        only_late = st.toggle("Uniquement en retard", value=False, key="contrib_only_late")

    filtered = sort_status_for_entry(filter_members_status(status_all, search, only_late))
    if filtered.empty:
        st.warning("Aucun membre ne correspond à votre recherche.")
        return

    st.caption(f"{len(filtered)} membre(s) affiché(s) — cliquez sur une ligne pour la sélectionner.")

    grid_df = status_grid_dataframe(filtered)
    id_by_row = filtered["id"].astype(int).tolist()

    selection = st.dataframe(
        grid_df,
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key="contrib_member_grid",
        column_config={
            "Référence": st.column_config.TextColumn(width="small"),
            "Statut": st.column_config.TextColumn(width="small"),
            "Reste": st.column_config.NumberColumn(format="%.2f EUR"),
            "Payé": st.column_config.NumberColumn(format="%.2f EUR"),
            "Attendu": st.column_config.NumberColumn(format="%.2f EUR"),
            "Solde N-1": st.column_config.NumberColumn(format="%.2f EUR"),
        },
    )

    selected_rows = []
    if selection is not None and hasattr(selection, "selection") and selection.selection is not None:
        selected_rows = list(selection.selection.rows or [])

    if selected_rows and 0 <= selected_rows[0] < len(id_by_row):
        st.session_state["contrib_member_id"] = id_by_row[selected_rows[0]]

    if "contrib_member_id" not in st.session_state or st.session_state["contrib_member_id"] not in id_by_row:
        st.session_state["contrib_member_id"] = id_by_row[0]

    current_id = int(st.session_state["contrib_member_id"])
    pick_options = {member_pick_label(r): int(r["id"]) for _, r in filtered.iterrows()}
    pick_labels = list(pick_options.keys())
    default_pick_idx = next(
        (i for i, lbl in enumerate(pick_labels) if pick_options[lbl] == current_id),
        0,
    )
    picked = st.selectbox(
        "Membre sélectionné",
        pick_labels,
        index=default_pick_idx,
        help="Alternative au clic dans le tableau — pratique sur téléphone.",
    )
    st.session_state["contrib_member_id"] = pick_options[picked]
    selected_member_id = int(st.session_state["contrib_member_id"])

    member_row = filtered.loc[filtered["id"] == selected_member_id].iloc[0]
    reste = float(member_row["reste"])
    attendu = float(member_row["attendu"])
    paye = float(member_row["total_paye"])
    solde_n1 = float(member_row["solde_n1"])
    statut_txt = str(member_row["statut"])
    statut_badge = "✅ À jour" if statut_txt == "A jour" else "⏳ En retard"
    solde_n1_txt = (
        f"Solde N-1 : {format_eur(solde_n1)}"
        if solde_n1 >= -0.001
        else f"Avance N-1 : {format_eur(-solde_n1)}"
    )

    st.divider()

    col_table, col_form = st.columns([1.45, 1], gap="large")

    with col_form:
        with st.container(border=True):
            st.markdown(f"#### {member_row['reference']} — {member_row['nom']} {member_row['prenom']}")
            st.caption(
                f"{statut_badge} · Village : {member_row['village_origine'] or '—'} · "
                f"{solde_n1_txt}"
            )

            m_a, m_b, m_c = st.columns(3)
            m_a.metric("Payé", format_eur(paye))
            m_b.metric("Attendu", format_eur(attendu))
            if reste < -0.001:
                m_c.metric("Avance", format_eur(-reste))
            else:
                m_c.metric("Reste", format_eur(reste))

            st.markdown("##### Enregistrement rapide")
            quick_mode = st.radio(
                "Mode de paiement",
                PAYMENT_METHODS,
                horizontal=True,
                key="contrib_quick_mode",
            )
            quick_col1, quick_col2 = st.columns(2)
            with quick_col1:
                if st.button(
                    f"+ {monthly_amount:.0f} EUR (aujourd'hui)",
                    type="primary",
                    use_container_width=True,
                    key="contrib_quick_month",
                ):
                    insert_contribution(
                        conn,
                        selected_member_id,
                        monthly_amount,
                        date.today(),
                        "Cotisation mensuelle",
                        quick_mode,
                    )
                    st.toast(f"Cotisation enregistrée pour {member_row['reference']}.")
                    st.rerun()
            with quick_col2:
                if reste > 0.001 and st.button(
                    f"Régler le reste ({format_eur(reste)})",
                    use_container_width=True,
                    key="contrib_quick_reste",
                ):
                    insert_contribution(
                        conn,
                        selected_member_id,
                        reste,
                        date.today(),
                        "Règlement du solde",
                        quick_mode,
                    )
                    st.toast(f"Solde réglé pour {member_row['reference']}.")
                    st.rerun()

            with st.expander("Régulariser plusieurs mois", expanded=False):
                st.caption(
                    "Sélectionnez les mois à payer. Chaque cotisation est datée au "
                    "**1er dimanche** (jour de réunion) et vaut le montant mensuel."
                )
                paid_months = {
                    int(m)
                    for (m,) in conn.execute(
                        """
                        SELECT DISTINCT CAST(strftime('%m', date) AS INTEGER)
                        FROM contributions
                        WHERE membre_id = ? AND strftime('%Y', date) = ?
                        """,
                        (selected_member_id, str(year_int)),
                    ).fetchall()
                }

                def _month_label(m: int) -> str:
                    mark = " ✅" if m in paid_months else ""
                    return f"{MONTHS_FR[m - 1]} ({first_sunday(year_int, m).strftime('%d/%m')}){mark}"

                month_choices = {_month_label(m): m for m in range(1, 13)}
                pick_nonce = st.session_state.get("contrib_months_nonce", 0)
                picked_months = st.multiselect(
                    "Mois à régulariser",
                    list(month_choices.keys()),
                    help="✅ = au moins une cotisation déjà enregistrée ce mois.",
                    key=f"contrib_months_pick_{pick_nonce}",
                )
                amount_per_month = st.number_input(
                    "Montant par mois (EUR)",
                    min_value=0.01,
                    value=float(monthly_amount),
                    step=1.0,
                    key="contrib_months_amount",
                )
                mode_m = st.radio(
                    "Mode de paiement",
                    PAYMENT_METHODS,
                    horizontal=True,
                    key="contrib_months_mode",
                )
                force_dup_m = st.checkbox(
                    "Autoriser les mois déjà réglés (doublons)",
                    value=False,
                    key="contrib_months_force",
                )
                nb_sel = len(picked_months)
                st.caption(
                    f"{nb_sel} mois sélectionné(s) — total : "
                    f"{format_eur(nb_sel * float(amount_per_month))}"
                )
                if st.button(
                    "Enregistrer les mois sélectionnés",
                    type="primary",
                    use_container_width=True,
                    disabled=nb_sel == 0,
                    key="contrib_months_submit",
                ):
                    chosen = [month_choices[lbl] for lbl in picked_months]
                    to_insert = [m for m in chosen if force_dup_m or m not in paid_months]
                    skipped = [m for m in chosen if not force_dup_m and m in paid_months]
                    for m in to_insert:
                        reunion = first_sunday(year_int, m)
                        insert_contribution(
                            conn,
                            selected_member_id,
                            float(amount_per_month),
                            reunion,
                            f"Cotisation {MONTHS_FR[m - 1]} {year_int}",
                            mode_m,
                        )
                    if skipped:
                        noms = ", ".join(MONTHS_FR[m - 1] for m in skipped)
                        st.warning(
                            f"{len(skipped)} mois déjà réglé(s) ignoré(s) : {noms}. "
                            "Cochez « Autoriser les doublons » pour forcer."
                        )
                    if to_insert:
                        st.session_state["contrib_months_nonce"] = pick_nonce + 1
                        st.toast(
                            f"{len(to_insert)} mois enregistré(s) pour {member_row['reference']}."
                        )
                        st.rerun()

            with st.expander("Autre montant ou date", expanded=False):
                with st.form("add_contribution_custom", clear_on_submit=True):
                    montant = st.number_input(
                        "Montant (EUR)",
                        min_value=0.01,
                        value=float(monthly_amount),
                        step=1.0,
                    )
                    contribution_date = st.date_input("Date", value=date.today())
                    mode_c = st.radio(
                        "Mode de paiement",
                        PAYMENT_METHODS,
                        horizontal=True,
                        key="contrib_custom_mode",
                    )
                    note = st.text_input("Note (facultatif)", placeholder="Ex. chèque, mobile money…")
                    force_dup_c = st.checkbox(
                        "Enregistrer même si ce mois est déjà réglé",
                        value=False,
                        key="contrib_custom_force",
                    )
                    if st.form_submit_button("Enregistrer", use_container_width=True):
                        if contribution_date.month in paid_months and not force_dup_c:
                            st.warning(
                                f"{MONTHS_FR[contribution_date.month - 1]} {year_int} a déjà une "
                                "cotisation. Cochez la case pour confirmer le doublon."
                            )
                        else:
                            insert_contribution(
                                conn,
                                selected_member_id,
                                float(montant),
                                contribution_date,
                                note,
                                mode_c,
                            )
                            st.success("Cotisation enregistrée.")
                            st.rerun()

    with col_table:
        st.markdown("##### Historique du membre")
        hist = fetch_df(
            conn,
            """
            SELECT c.id, c.date, c.montant, c.mode_paiement, c.note
            FROM contributions c
            WHERE strftime('%Y', c.date) = ?
              AND c.membre_id = ?
            ORDER BY c.date DESC, c.id DESC;
            """,
            (str(year_int), selected_member_id),
        )
        if hist.empty:
            st.info("Aucune cotisation enregistrée pour cet exercice.")
        else:
            st.dataframe(
                hist,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "date": st.column_config.TextColumn("Date"),
                    "montant": st.column_config.NumberColumn("Montant", format="%.2f EUR"),
                    "mode_paiement": st.column_config.TextColumn("Mode"),
                    "note": st.column_config.TextColumn("Note"),
                },
            )

            with st.expander("Corriger ou supprimer une ligne", expanded=False):
                row_options = {
                    f"{r['date']} · {float(r['montant']):.2f} EUR · {(r['note'] or '—')}": int(r["id"])
                    for _, r in hist.iterrows()
                }
                selected_label = st.selectbox(
                    "Ligne à corriger",
                    list(row_options.keys()),
                    key="contrib_edit_pick",
                )
                selected_id = row_options[selected_label]
                row = conn.execute(
                    "SELECT id, membre_id, montant, date, note, mode_paiement FROM contributions WHERE id = ?",
                    (selected_id,),
                ).fetchone()
                if row:
                    e1, e2 = st.columns(2)
                    with e1:
                        edit_amount = st.number_input(
                            "Montant",
                            min_value=0.01,
                            value=float(row["montant"]),
                            step=1.0,
                            key=f"edit_amount_{selected_id}",
                        )
                        edit_date = st.date_input(
                            "Date",
                            value=datetime.fromisoformat(row["date"]).date(),
                            key=f"edit_date_{selected_id}",
                        )
                    with e2:
                        current_mode = row["mode_paiement"] or DEFAULT_PAYMENT_METHOD
                        edit_mode = st.radio(
                            "Mode de paiement",
                            PAYMENT_METHODS,
                            index=PAYMENT_METHODS.index(current_mode)
                            if current_mode in PAYMENT_METHODS
                            else 0,
                            horizontal=True,
                            key=f"edit_mode_{selected_id}",
                        )
                        edit_note = st.text_input(
                            "Note",
                            value=row["note"] or "",
                            key=f"edit_note_{selected_id}",
                        )
                    b1, b2 = st.columns(2)
                    with b1:
                        if st.button("Mettre à jour", key=f"update_contrib_{selected_id}"):
                            before = fmt_contribution_compact(
                                float(row["montant"]), str(row["date"]), row["note"] or ""
                            ) + f" [{row['mode_paiement'] or DEFAULT_PAYMENT_METHOD}]"
                            conn.execute(
                                """
                                UPDATE contributions
                                SET montant = ?, date = ?, note = ?, mode_paiement = ?
                                WHERE id = ?
                                """,
                                (
                                    float(edit_amount),
                                    to_iso(edit_date),
                                    edit_note.strip(),
                                    edit_mode,
                                    selected_id,
                                ),
                            )
                            after = fmt_contribution_compact(
                                float(edit_amount), to_iso(edit_date), edit_note.strip()
                            ) + f" [{edit_mode}]"
                            log_activity(
                                conn,
                                type_operation="UPDATE",
                                entite="contribution",
                                entite_id=selected_id,
                                details=f"Maj cotisation #{selected_id} {before}  ->  {after}",
                            )
                            conn.commit()
                            st.success("Cotisation mise à jour.")
                            st.rerun()
                    with b2:
                        if st.button("Supprimer", type="secondary", key=f"delete_contrib_{selected_id}"):
                            conn.execute("DELETE FROM contributions WHERE id = ?", (selected_id,))
                            log_activity(
                                conn,
                                type_operation="DELETE",
                                entite="contribution",
                                entite_id=selected_id,
                                details=f"Suppression cotisation #{selected_id}",
                            )
                            conn.commit()
                            st.success("Cotisation supprimée.")
                            st.rerun()

            with st.expander("🧾 Reçu de paiement (PDF)", expanded=False):
                receipt_options = {
                    f"{r['date']} · {float(r['montant']):.2f} EUR · {(r['note'] or 'Cotisation')}": (
                        int(r["id"]),
                        str(r["date"]),
                        float(r["montant"]),
                        r["note"] or "",
                        r["mode_paiement"] or DEFAULT_PAYMENT_METHOD,
                    )
                    for _, r in hist.iterrows()
                }
                receipt_label = st.selectbox(
                    "Cotisation à justifier",
                    list(receipt_options.keys()),
                    key="receipt_pick",
                )
                _, r_date, r_montant, r_note, r_mode = receipt_options[receipt_label]
                pdf_bytes = build_receipt_pdf(
                    reference=str(member_row["reference"]),
                    nom=str(member_row["nom"]),
                    prenom=str(member_row["prenom"]),
                    montant=r_montant,
                    date_paiement=r_date,
                    note=r_note,
                    mode_paiement=r_mode,
                    annee=year_int,
                    total_paye_annee=paye,
                    reste_annee=reste,
                )
                if pdf_bytes is None:
                    st.warning(
                        "Le module de génération PDF n'est pas installé. "
                        "Ajoutez `fpdf2` aux dépendances (requirements.txt)."
                    )
                else:
                    st.download_button(
                        "⬇️ Télécharger le reçu",
                        data=pdf_bytes,
                        file_name=f"recu_{member_row['reference']}_{r_date}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="receipt_download",
                    )

    with st.expander("Voir toutes les cotisations de l'exercice", expanded=False):
        all_hist = fetch_df(
            conn,
            """
            SELECT c.date, m.reference, m.nom, m.prenom, c.montant, c.mode_paiement, c.note
            FROM contributions c
            JOIN membres m ON m.id = c.membre_id
            WHERE strftime('%Y', c.date) = ?
            ORDER BY c.date DESC, c.id DESC;
            """,
            (str(year_int),),
        )
        hist_search = st.text_input(
            "Filtrer l'historique global",
            "",
            key="contrib_hist_global_search",
            placeholder="Nom, référence, note…",
        ).strip().lower()
        if hist_search and not all_hist.empty:
            mask = (
                all_hist["reference"].astype(str).str.lower().str.contains(hist_search, na=False)
                | all_hist["nom"].astype(str).str.lower().str.contains(hist_search, na=False)
                | all_hist["prenom"].astype(str).str.lower().str.contains(hist_search, na=False)
                | all_hist["note"].astype(str).str.lower().str.contains(hist_search, na=False)
            )
            all_hist = all_hist[mask]
        st.dataframe(all_hist, use_container_width=True, hide_index=True)
        st.caption(
            f"Attendu par membre = {monthly_amount:.0f} EUR/mois (selon inscription) "
            f"+ solde N-1 reporté · Total encaissé : {format_eur(total_year)}"
        )
        st.divider()
        st.markdown("**Exports pour archives / réunions**")
        ce1, ce2 = st.columns(2)
        with ce1:
            st.caption("Cotisations de l'exercice")
            render_export_buttons(all_hist, f"cotisations_{year_int}", key="export_contribs")
        with ce2:
            st.caption("Statut des membres")
            render_export_buttons(
                status_display_df(status_all), f"statut_membres_{year_int}", key="export_status"
            )


def page_depenses(conn: sqlite3.Connection) -> None:
    st.subheader("Dépenses")
    render_page_guide(
        "1️⃣ Choisissez l'exercice &nbsp;→&nbsp; 2️⃣ Enregistrez une dépense à droite "
        "ou sélectionnez une ligne pour la corriger."
    )

    year = st.number_input(
        "Exercice",
        min_value=2020,
        max_value=2100,
        value=date.today().year,
        step=1,
        key="depenses_year",
    )
    year_int = int(year)
    total_dep = total_expenses(conn, year_int)
    nb_dep = conn.execute(
        "SELECT COUNT(*) AS n FROM depenses WHERE strftime('%Y', date) = ?",
        (str(year_int),),
    ).fetchone()["n"]

    k1, k2 = st.columns(2)
    k1.metric("Dépenses enregistrées", int(nb_dep or 0))
    k2.metric("Total (année)", format_eur(total_dep))

    dep = fetch_df(
        conn,
        """
        SELECT id, date, description, montant
        FROM depenses
        WHERE strftime('%Y', date) = ?
        ORDER BY date DESC, id DESC;
        """,
        (str(year_int),),
    )

    search = st.text_input(
        "Rechercher",
        "",
        key="depenses_search",
        placeholder="Description, date…",
    )
    dep = filter_df_search(dep, search, ["description", "date"])
    selected_dep_id: Optional[int] = None

    col_list, col_form = st.columns([1.5, 1], gap="large")

    with col_list:
        st.markdown("##### Liste des dépenses")
        if dep.empty:
            st.info("Aucune dépense pour cet exercice.")
            selected_dep_id = None
        else:
            st.caption(f"{len(dep)} ligne(s) — cliquez pour sélectionner.")
            grid = dep.rename(
                columns={"date": "Date", "description": "Description", "montant": "Montant (EUR)"}
            ).drop(columns=["id"], errors="ignore")
            id_by_row = dep["id"].astype(int).tolist()

            selection = st.dataframe(
                grid,
                use_container_width=True,
                hide_index=True,
                on_select="rerun",
                selection_mode="single-row",
                key="depenses_grid",
                column_config={
                    "Montant (EUR)": st.column_config.NumberColumn(format="%.2f EUR"),
                },
            )
            render_export_buttons(grid, f"depenses_{year_int}", key="export_depenses")
            selected_rows = []
            if selection is not None and hasattr(selection, "selection") and selection.selection:
                selected_rows = list(selection.selection.rows or [])

            if selected_rows and 0 <= selected_rows[0] < len(id_by_row):
                st.session_state["depense_selected_id"] = id_by_row[selected_rows[0]]
            elif id_by_row:
                if st.session_state.get("depense_selected_id") not in id_by_row:
                    st.session_state["depense_selected_id"] = id_by_row[0]

            dep_options = {
                f"{r['date']} · {r['description']} · {float(r['montant']):.2f} EUR": int(r["id"])
                for _, r in dep.iterrows()
            }
            dep_labels = list(dep_options.keys())
            current_dep = st.session_state.get("depense_selected_id", id_by_row[0] if id_by_row else None)
            if current_dep is not None and dep_labels:
                dep_idx = next(
                    (i for i, lbl in enumerate(dep_labels) if dep_options[lbl] == current_dep),
                    0,
                )
                picked_dep = st.selectbox("Ligne sélectionnée", dep_labels, index=dep_idx)
                selected_dep_id = dep_options[picked_dep]
                st.session_state["depense_selected_id"] = selected_dep_id
            else:
                selected_dep_id = None

    with col_form:
        with st.container(border=True):
            st.markdown("##### Nouvelle dépense")
            with st.form("add_expense", clear_on_submit=True):
                description = st.text_input("Description *", placeholder="Ex. location salle, fournitures…")
                montant = st.number_input("Montant (EUR)", min_value=0.01, value=10.0, step=1.0)
                depense_date = st.date_input("Date", value=date.today())
                if st.form_submit_button("Enregistrer", type="primary", use_container_width=True):
                    if not description.strip():
                        st.error("Description obligatoire.")
                    else:
                        insert_depense(conn, description, float(montant), depense_date)
                        st.toast("Dépense enregistrée.")
                        st.rerun()

        if dep.empty:
            pass
        elif selected_dep_id is not None:
            with st.container(border=True):
                st.markdown("##### Modifier la sélection")
                dep_row = conn.execute(
                    "SELECT id, date, description, montant FROM depenses WHERE id = ?",
                    (selected_dep_id,),
                ).fetchone()
                if dep_row:
                    edit_dep_desc = st.text_input(
                        "Description",
                        value=dep_row["description"] or "",
                        key=f"edit_dep_desc_{selected_dep_id}",
                    )
                    edit_dep_amount = st.number_input(
                        "Montant (EUR)",
                        min_value=0.01,
                        value=float(dep_row["montant"]),
                        step=1.0,
                        key=f"edit_dep_amount_{selected_dep_id}",
                    )
                    edit_dep_date = st.date_input(
                        "Date",
                        value=datetime.fromisoformat(dep_row["date"]).date(),
                        key=f"edit_dep_date_{selected_dep_id}",
                    )
                    x1, x2 = st.columns(2)
                    with x1:
                        if st.button("Mettre à jour", key=f"update_dep_{selected_dep_id}", use_container_width=True):
                            if not edit_dep_desc.strip():
                                st.error("Description obligatoire.")
                            else:
                                before = fmt_depense_compact(
                                    dep_row["description"] or "",
                                    float(dep_row["montant"]),
                                    str(dep_row["date"]),
                                )
                                conn.execute(
                                    """
                                    UPDATE depenses
                                    SET description = ?, montant = ?, date = ?
                                    WHERE id = ?
                                    """,
                                    (
                                        edit_dep_desc.strip(),
                                        float(edit_dep_amount),
                                        to_iso(edit_dep_date),
                                        selected_dep_id,
                                    ),
                                )
                                after = fmt_depense_compact(
                                    edit_dep_desc.strip(),
                                    float(edit_dep_amount),
                                    to_iso(edit_dep_date),
                                )
                                log_activity(
                                    conn,
                                    type_operation="UPDATE",
                                    entite="depense",
                                    entite_id=selected_dep_id,
                                    details=f"Maj depense #{selected_dep_id} {before}  ->  {after}",
                                )
                                conn.commit()
                                st.rerun()
                    with x2:
                        if st.button(
                            "Supprimer",
                            type="secondary",
                            key=f"delete_dep_{selected_dep_id}",
                            use_container_width=True,
                        ):
                            conn.execute("DELETE FROM depenses WHERE id = ?", (selected_dep_id,))
                            log_activity(
                                conn,
                                type_operation="DELETE",
                                entite="depense",
                                entite_id=selected_dep_id,
                                details=f"Suppression depense #{selected_dep_id}",
                            )
                            conn.commit()
                            st.rerun()


def page_dashboard(conn: sqlite3.Connection) -> None:
    st.subheader("Tableau de bord")
    render_page_guide("Vue d'ensemble de l'exercice : encaissements, dépenses et situation des membres.")

    head_l, head_r = st.columns([3, 1])
    with head_r:
        year = int(
            st.number_input(
                "Exercice",
                min_value=2020,
                max_value=2100,
                value=date.today().year,
                step=1,
                key="dashboard_year",
            )
        )
    with head_l:
        st.markdown(f"#### Exercice {year}")

    contrib = total_contributions(conn, year)
    dep = total_expenses(conn, year)
    report = get_association_report(conn, year)
    report_n2 = get_association_report(conn, year - 1)
    solde = report + contrib - dep

    contrib_prev = total_contributions(conn, year - 1)
    dep_prev = total_expenses(conn, year - 1)
    solde_prev = report_n2 + contrib_prev - dep_prev

    st.divider()

    status_year = get_members_status(conn, year)
    nb_late = int((status_year["statut"] == "En retard").sum()) if not status_year.empty else 0
    nb_ok = len(status_year) - nb_late if not status_year.empty else 0

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric(
        "Cotisations",
        format_eur(contrib),
        delta=f"{(contrib - contrib_prev):+,.2f} EUR vs N-1".replace(",", " "),
    )
    k2.metric(
        "Dépenses",
        format_eur(dep),
        delta=f"{(dep - dep_prev):+,.2f} EUR vs N-1".replace(",", " "),
        delta_color="inverse",
    )
    k3.metric("Report N-1", format_eur(report), help=f"Solde N-2 : {format_eur(report_n2)}")
    k4.metric(
        "Solde actuel",
        format_eur(solde),
        delta=f"{(solde - solde_prev):+,.2f} EUR vs N-1".replace(",", " "),
    )
    k5.metric("Membres à jour", nb_ok)
    k6.metric("Membres en retard", nb_late)

    st.divider()

    tab_apercu, tab_membres, tab_evolution, tab_parametres = st.tabs(
        ["Vue d'ensemble", "Membres", "Évolution mensuelle", "Paramètres"]
    )

    with tab_apercu:
        st.markdown("##### Du report N-1 au solde actuel")

        steps = [
            ("Report N-1", report, "Report"),
            ("Contributions", contrib, "Apport"),
            ("Dépenses", -dep, "Sortie"),
            ("Solde actuel", solde, "Solde"),
        ]
        wf = pd.DataFrame(steps, columns=["etape", "delta", "type"])

        running = 0.0
        bases: list[float] = []
        tops: list[float] = []
        for _, r in wf.iterrows():
            if r["type"] == "Solde":
                bases.append(0.0)
                tops.append(float(r["delta"]))
            else:
                start = running
                running += float(r["delta"])
                bases.append(min(start, running))
                tops.append(max(start, running))
        wf["base"] = bases
        wf["top"] = tops

        order = ["Report N-1", "Contributions", "Dépenses", "Solde actuel"]
        color_scale_wf = alt.Scale(
            domain=["Report", "Apport", "Sortie", "Solde"],
            range=["#94a3b8", "#16a34a", "#ef4444", "#2563eb"],
        )
        base_wf = alt.Chart(wf).encode(
            x=alt.X("etape:N", sort=order, title=None, axis=alt.Axis(labelAngle=0)),
        )
        bars_wf = base_wf.mark_bar(cornerRadius=5, size=58).encode(
            y=alt.Y("base:Q", title="Montant (EUR)", axis=alt.Axis(format=",.0f")),
            y2="top:Q",
            color=alt.Color("type:N", scale=color_scale_wf, legend=None),
            tooltip=[
                alt.Tooltip("etape:N", title="Étape"),
                alt.Tooltip("delta:Q", title="Montant", format="+,.2f"),
            ],
        )
        labels_wf = base_wf.mark_text(dy=-8, color="#1f2937", fontWeight="bold").encode(
            y="top:Q",
            text=alt.Text("delta:Q", format="+,.0f"),
        )
        chart_wf = (
            (bars_wf + labels_wf)
            .properties(height=340)
            .configure_view(strokeWidth=0)
            .configure_axis(grid=True, gridOpacity=0.2)
        )
        st.altair_chart(chart_wf, use_container_width=True)
        st.caption(
            "Lecture de gauche à droite : on part du report N-1, on ajoute les cotisations (vert), "
            "on retire les dépenses (rouge) → solde actuel (bleu)."
        )

        st.markdown("##### Synthèse N / N-1")
        synth = pd.DataFrame(
            [
                {
                    "Indicateur": "Contributions",
                    f"Année {year}": f"{contrib:,.2f} EUR".replace(",", " "),
                    f"Année {year - 1}": f"{contrib_prev:,.2f} EUR".replace(",", " "),
                    "Variation": f"{(contrib - contrib_prev):+,.2f} EUR".replace(",", " "),
                },
                {
                    "Indicateur": "Dépenses",
                    f"Année {year}": f"{dep:,.2f} EUR".replace(",", " "),
                    f"Année {year - 1}": f"{dep_prev:,.2f} EUR".replace(",", " "),
                    "Variation": f"{(dep - dep_prev):+,.2f} EUR".replace(",", " "),
                },
                {
                    "Indicateur": "Solde",
                    f"Année {year}": f"{solde:,.2f} EUR".replace(",", " "),
                    f"Année {year - 1}": f"{solde_prev:,.2f} EUR".replace(",", " "),
                    "Variation": f"{(solde - solde_prev):+,.2f} EUR".replace(",", " "),
                },
            ]
        )
        st.dataframe(synth, use_container_width=True, hide_index=True)

    with tab_membres:
        st.markdown("##### Situation des cotisations")
        if status_year.empty:
            st.info("Aucun membre actif.")
        else:
            total_membres = len(status_year)
            taux = (nb_ok / total_membres * 100) if total_membres else 0.0
            attendu_total = float(status_year["attendu"].clip(lower=0).sum())
            reste_total = float(status_year["reste"].clip(lower=0).sum())
            credit_total = float((-status_year["reste"]).clip(lower=0).sum())

            mk1, mk2, mk3 = st.columns(3)
            mk1.metric("Membres à jour", f"{nb_ok} / {total_membres}", help=f"{taux:.0f}% des membres actifs")
            mk2.metric("En retard", nb_late)
            credit_help = (
                f"Attendu total : {format_eur(attendu_total)}"
                + (f" · Avances : {format_eur(credit_total)}" if credit_total > 0.001 else "")
            )
            mk3.metric("Reste à encaisser", format_eur(reste_total), help=credit_help)

            if nb_late > 0:
                st.warning(
                    f"⏳ {nb_late} membre(s) en retard sur l'exercice {year} "
                    f"— détail et encaissement dans le menu **Cotisations**."
                )
            else:
                st.success(f"✅ Tous les membres sont à jour pour l'exercice {year}.")

    with tab_evolution:
        st.markdown("##### Contributions et dépenses par mois")
        monthly = fetch_df(
            conn,
            """
            WITH months AS (
                SELECT '01' m UNION ALL SELECT '02' UNION ALL SELECT '03' UNION ALL SELECT '04'
                UNION ALL SELECT '05' UNION ALL SELECT '06' UNION ALL SELECT '07' UNION ALL SELECT '08'
                UNION ALL SELECT '09' UNION ALL SELECT '10' UNION ALL SELECT '11' UNION ALL SELECT '12'
            ),
            c AS (
                SELECT strftime('%m', date) m, SUM(montant) total
                FROM contributions
                WHERE strftime('%Y', date) = ?
                GROUP BY strftime('%m', date)
            ),
            d AS (
                SELECT strftime('%m', date) m, SUM(montant) total
                FROM depenses
                WHERE strftime('%Y', date) = ?
                GROUP BY strftime('%m', date)
            )
            SELECT months.m AS mois,
                   COALESCE(c.total, 0) AS contributions,
                   COALESCE(d.total, 0) AS depenses
            FROM months
            LEFT JOIN c ON c.m = months.m
            LEFT JOIN d ON d.m = months.m
            ORDER BY months.m;
            """,
            (str(year), str(year)),
        )
        monthly["solde_net"] = monthly["contributions"] - monthly["depenses"]
        monthly["cumul_contributions"] = monthly["contributions"].cumsum()
        monthly["cumul_depenses"] = monthly["depenses"].cumsum()
        monthly["cumul_solde"] = monthly["cumul_contributions"] - monthly["cumul_depenses"]

        month_labels_short = [
            "Jan", "Fév", "Mar", "Avr", "Mai", "Juin",
            "Juil", "Août", "Sep", "Oct", "Nov", "Déc",
        ]
        monthly["mois_label"] = monthly["mois"].apply(
            lambda m: month_labels_short[int(m) - 1] if pd.notna(m) else ""
        )
        month_order = month_labels_short

        flux_df = monthly.melt(
            id_vars=["mois_label"],
            value_vars=["contributions", "depenses"],
            var_name="serie",
            value_name="montant",
        )
        flux_df["serie"] = flux_df["serie"].map(
            {"contributions": "Contributions", "depenses": "Dépenses"}
        )
        color_flux = alt.Scale(
            domain=["Contributions", "Dépenses"], range=["#16a34a", "#ef4444"]
        )

        bars = (
            alt.Chart(flux_df)
            .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
            .encode(
                x=alt.X(
                    "mois_label:N",
                    sort=month_order,
                    title="Mois",
                    axis=alt.Axis(labelAngle=0),
                ),
                xOffset=alt.XOffset("serie:N", sort=["Contributions", "Dépenses"]),
                y=alt.Y("montant:Q", title="Montant (EUR)", axis=alt.Axis(format=",.0f")),
                color=alt.Color(
                    "serie:N",
                    scale=color_flux,
                    legend=alt.Legend(title=None, orient="top"),
                ),
                tooltip=[
                    alt.Tooltip("mois_label:N", title="Mois"),
                    alt.Tooltip("serie:N", title="Type"),
                    alt.Tooltip("montant:Q", title="Montant", format=",.2f"),
                ],
            )
        )
        line_solde = (
            alt.Chart(monthly)
            .mark_line(color="#2563eb", strokeWidth=2.5, point=alt.OverlayMarkDef(filled=True, size=70))
            .encode(
                x=alt.X("mois_label:N", sort=month_order),
                y=alt.Y("solde_net:Q"),
                tooltip=[
                    alt.Tooltip("mois_label:N", title="Mois"),
                    alt.Tooltip("solde_net:Q", title="Solde net", format=",.2f"),
                ],
            )
        )
        chart_flux = (
            (bars + line_solde)
            .properties(height=320)
            .configure_view(strokeWidth=0)
            .configure_axis(grid=True, gridOpacity=0.25)
        )
        st.altair_chart(chart_flux, use_container_width=True)
        st.caption("Barres : flux mensuels — Ligne bleue : solde net (contributions − dépenses).")

        st.markdown("##### Cumul depuis janvier")
        cumul_df = monthly.melt(
            id_vars=["mois_label"],
            value_vars=["cumul_contributions", "cumul_depenses", "cumul_solde"],
            var_name="serie",
            value_name="montant",
        )
        cumul_df["serie"] = cumul_df["serie"].map(
            {
                "cumul_contributions": "Contributions cumulées",
                "cumul_depenses": "Dépenses cumulées",
                "cumul_solde": "Solde cumulé",
            }
        )
        color_cumul = alt.Scale(
            domain=["Contributions cumulées", "Dépenses cumulées", "Solde cumulé"],
            range=["#16a34a", "#ef4444", "#2563eb"],
        )
        chart_cumul = (
            alt.Chart(cumul_df)
            .mark_area(opacity=0.35, line={"strokeWidth": 2})
            .encode(
                x=alt.X(
                    "mois_label:N",
                    sort=month_order,
                    title="Mois",
                    axis=alt.Axis(labelAngle=0),
                ),
                y=alt.Y("montant:Q", title="Montant cumulé (EUR)", axis=alt.Axis(format=",.0f"), stack=None),
                color=alt.Color(
                    "serie:N",
                    scale=color_cumul,
                    legend=alt.Legend(title=None, orient="top"),
                ),
                tooltip=[
                    alt.Tooltip("mois_label:N", title="Mois"),
                    alt.Tooltip("serie:N", title="Série"),
                    alt.Tooltip("montant:Q", title="Cumul", format=",.2f"),
                ],
            )
            .properties(height=280)
            .configure_view(strokeWidth=0)
            .configure_axis(grid=True, gridOpacity=0.25)
        )
        st.altair_chart(chart_cumul, use_container_width=True)

    with tab_parametres:
        col_p1, col_p2 = st.columns(2)

        with col_p1:
            with st.container(border=True):
                st.markdown("##### Montant de la cotisation mensuelle")
                st.caption(
                    "Montant attendu par membre et par mois. Sert au calcul du statut "
                    "et aux boutons d'enregistrement rapide."
                )
                current_amount = get_monthly_contribution(conn)
                new_amount = st.number_input(
                    "Montant mensuel (EUR)",
                    min_value=0.01,
                    value=float(current_amount),
                    step=1.0,
                    format="%.2f",
                    key="settings_monthly_amount",
                )
                if st.button("Enregistrer le montant", type="primary", key="save_monthly_amount"):
                    set_monthly_contribution(conn, float(new_amount))
                    st.success(f"Cotisation mensuelle fixée à {format_eur(float(new_amount))}.")
                    st.rerun()
                st.caption(
                    "ℹ️ Modifier le montant n'altère pas les cotisations déjà enregistrées ; "
                    "cela change l'attendu pour le calcul des statuts."
                )

        with col_p2:
            with st.container(border=True):
                st.markdown("##### Solde reporté de l'association")
                st.caption(
                    "Solde repris de l'année précédente vers l'exercice en cours. "
                    "Alimente le KPI « Report N-1 » et le solde global."
                )
                new_report = st.number_input(
                    f"Solde reporté (N-1 → {year})",
                    value=float(report),
                    step=10.0,
                    key=f"dashboard_report_{year}",
                )
                if st.button("Enregistrer le solde reporté", type="primary"):
                    upsert_association_report(conn, year, float(new_report))
                    st.success("Solde reporté association enregistré.")
                    st.rerun()

        with col_p1:
            with st.container(border=True):
                st.markdown("##### Indicatif téléphonique (rappels)")
                st.caption(
                    "Utilisé pour générer les liens WhatsApp des retardataires. "
                    "Ex. 33 pour la France, 224 pour la Guinée."
                )
                current_cc = get_default_country_code(conn)
                new_cc = st.text_input(
                    "Indicatif pays (sans +)",
                    value=current_cc,
                    key="settings_country_code",
                )
                if st.button("Enregistrer l'indicatif", key="save_country_code"):
                    set_default_country_code(conn, new_cc)
                    st.success(f"Indicatif fixé à +{get_default_country_code(conn)}.")
                    st.rerun()

        with col_p2:
            with st.container(border=True):
                base_year = get_baseline_year(conn)
                st.markdown("##### Reports obsolètes")
                st.caption(
                    f"Le report N-1 est recalculé automatiquement depuis {base_year}. "
                    f"Les soldes figés postérieurs à {base_year} (ex. import) sont inutiles "
                    "et peuvent être supprimés."
                )
                nb_obsolete = conn.execute(
                    "SELECT COUNT(*) AS n FROM reports_membres WHERE annee > ?",
                    (base_year,),
                ).fetchone()["n"]
                st.metric("Reports obsolètes en base", int(nb_obsolete or 0))
                if st.button(
                    "Nettoyer les reports obsolètes",
                    type="primary",
                    disabled=int(nb_obsolete or 0) == 0,
                    key="purge_reports",
                ):
                    removed = conn.execute(
                        "DELETE FROM reports_membres WHERE annee > ?;", (base_year,)
                    ).rowcount
                    log_activity(
                        conn,
                        type_operation="DELETE",
                        entite="reports_membres",
                        entite_id=None,
                        details=f"Nettoyage manuel : {removed} report(s) postérieur(s) à {base_year}.",
                    )
                    conn.commit()
                    st.success(f"{removed} report(s) obsolète(s) supprimé(s).")
                    st.rerun()


def page_activite(conn: sqlite3.Connection) -> None:
    st.subheader("Journal d'activité")
    render_page_guide("Historique des actions : cotisations, dépenses, modifications de fiches.")

    f_lim, f_search, f_type = st.columns([1, 2, 1])
    with f_lim:
        limit = st.selectbox("Afficher les", [20, 50, 100, 200], index=1, key="activity_limit")
    with f_search:
        search = st.text_input(
            "Rechercher",
            "",
            key="activity_search",
            placeholder="Membre, référence, détail…",
        )
    with f_type:
        type_filter = st.selectbox(
            "Type",
            ["Tous", "Cotisation", "Dépense", "Membre", "Import"],
            key="activity_type_filter",
        )

    logs = fetch_df(
        conn,
        """
        SELECT a.id,
               a.created_at,
               a.entite,
               a.entite_id,
               COALESCE(
                   NULLIF(TRIM(mm.nom || ' ' || mm.prenom), ''),
                   NULLIF(TRIM(mc.nom || ' ' || mc.prenom), ''),
                   ''
               ) AS nom_prenom,
               a.details
        FROM activites a
        LEFT JOIN membres mm
               ON a.entite IN ('membre', 'reports_membres')
              AND mm.id = a.entite_id
        LEFT JOIN contributions c
               ON a.entite = 'contribution'
              AND c.id = a.entite_id
        LEFT JOIN membres mc
               ON mc.id = c.membre_id
        ORDER BY a.id DESC
        LIMIT ?;
        """,
        (int(limit),),
    )

    if not logs.empty:
        logs["Type"] = logs["entite"].map(activity_entite_label)
        type_map = {
            "Cotisation": "contribution",
            "Dépense": "depense",
            "Membre": "membre",
            "Import": "import_excel",
        }
        if type_filter != "Tous":
            ent = type_map.get(type_filter, "")
            logs = logs[logs["entite"] == ent]
        logs = filter_df_search(logs, search, ["nom_prenom", "details", "Type"])
        display = logs.rename(
            columns={
                "created_at": "Date",
                "Type": "Type",
                "nom_prenom": "Membre",
                "details": "Détail",
            }
        )[["Date", "Type", "Membre", "Détail"]]

    if logs.empty:
        st.info("Aucun mouvement pour ces critères.")
    else:
        st.caption(f"{len(display)} mouvement(s)")
        st.dataframe(display, use_container_width=True, hide_index=True)


# --- Import Excel (structure classeur AGPM Association) ---------------------------------

MONTH_ORDER_FR = [
    "janvier",
    "fevrier",
    "mars",
    "avril",
    "mai",
    "juin",
    "juillet",
    "aout",
    "septembre",
    "octobre",
    "novembre",
    "decembre",
]

SOLDE_YEAR_IN_COL_RE = re.compile(r"solde\s*(\d{4})", re.IGNORECASE)
SOLDE_N1_COL_RE = re.compile(r"^solde\s+n\s*[- ]?\s*1$", re.IGNORECASE)
COTISATIONS_SHEET_RE = re.compile(r"^cotisations\s*(\d{4})$", re.IGNORECASE)
DEPENSES_SHEET_RE = re.compile(r"^d[ée]penses\s*(\d{4})$", re.IGNORECASE)


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def normalize_header(name: object) -> str:
    if pd.isna(name):
        return ""
    return _strip_accents(str(name).strip().lower())


def normalize_phone(raw: object) -> str:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    return re.sub(r"\D", "", str(raw))


def excel_member_key(nom: object, prenom: object, tel_raw: object) -> tuple[str, str, str]:
    n = str(nom).strip().lower() if pd.notna(nom) else ""
    p = str(prenom).strip().lower() if pd.notna(prenom) else ""
    return (n, p, normalize_phone(tel_raw))


# UID import : ("t", nom_l, prenom_l, tel) si téléphone dès Excel ; sinon ("r", feuille, n° ligne) pour éviter
# de fusionner deux homonymes sans numéro avant que vous complétiez le téléphone dans l'interface.
MemberImportUid = tuple[str | int, ...]


def member_import_uid(nom: object, prenom: object, tel_raw: object, sheet: str, row_num: int) -> MemberImportUid:
    n = str(nom).strip().lower() if pd.notna(nom) else ""
    p = str(prenom).strip().lower() if pd.notna(prenom) else ""
    t = normalize_phone(tel_raw)
    if t:
        return ("t", n, p, t)
    return ("r", sheet, row_num)


def uid_import_to_json(uid: MemberImportUid) -> str:
    return json.dumps(list(uid))


def uid_import_from_json(s: str) -> MemberImportUid:
    L = json.loads(s)
    kind = L[0]
    if kind == "r":
        return ("r", str(L[1]), int(L[2]))
    return ("t", str(L[1]), str(L[2]), str(L[3]))


def member_import_row_label(uid: MemberImportUid) -> str:
    if uid[0] == "r":
        return f"{uid[1]}, ligne {uid[2]}"
    return "Téléphone lu depuis Excel"


def sync_member_phones_from_editor(bundle: dict[str, Any], edited: pd.DataFrame) -> None:
    """Met à jour bundle['members'][uid]['telephone'] depuis la colonne éditée (_uid_json + telephone)."""
    if edited.empty or "_uid_json" not in edited.columns:
        return
    members = bundle["members"]
    col_tel = "telephone" if "telephone" in edited.columns else None
    if not col_tel:
        return
    for _, row in edited.iterrows():
        uid = uid_import_from_json(str(row["_uid_json"]))
        if uid not in members:
            continue
        raw_tel = row[col_tel]
        members[uid]["telephone"] = normalize_phone(raw_tel)


def collapse_bundle_members_by_phone(bundle: dict[str, Any]) -> None:
    """Fusionne les entrées ayant le même nom+prénom+téléphone (téléphone non vide après édition)."""
    members: dict[MemberImportUid, dict[str, str]] = bundle["members"]
    contributions: list[dict[str, object]] = bundle["contributions"]
    reports: list[dict[str, object]] = bundle["reports"]

    groups: dict[tuple[str, str, str], list[MemberImportUid]] = {}
    for uid, info in members.items():
        dk = excel_member_key(info["nom"], info["prenom"], info["telephone"])
        if dk[2] == "":
            continue
        groups.setdefault(dk, []).append(uid)

    uid_redirect: dict[MemberImportUid, MemberImportUid] = {uid: uid for uid in members}

    for uids in groups.values():
        if len(uids) < 2:
            continue
        canon = uids[0]
        merged = dict(members[canon])
        for u in uids[1:]:
            ou = members[u]
            merged["telephone"] = merge_member_field(merged["telephone"], ou["telephone"])
            merged["email"] = merge_member_field(merged["email"], ou["email"])
            merged["adresse"] = merge_member_field(merged["adresse"], ou["adresse"])
            merged["village_origine"] = merge_member_field(merged["village_origine"], ou["village_origine"])
            merged["prefecture"] = merge_member_field(merged["prefecture"], ou["prefecture"])
            uid_redirect[u] = canon
            del members[u]
        members[canon] = merged

    for c in contributions:
        ouid = c["member_uid"]
        c["member_uid"] = uid_redirect.get(ouid, ouid)
    for r in reports:
        ouid = r["member_uid"]
        r["member_uid"] = uid_redirect.get(ouid, ouid)


def validate_bundle_phones(bundle: dict[str, Any]) -> list[str]:
    errs: list[str] = []
    for uid, info in bundle["members"].items():
        if not normalize_phone(info["telephone"]):
            label = f"{info['nom']} {info['prenom']}"
            if uid[0] == "r":
                label += f" ({uid[1]}, ligne {uid[2]})"
            errs.append(label)
    return errs


def parse_money_cell(val: object) -> Optional[float]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, str):
        v = val.strip().lower()
        if v in ("", "x", "p", "a", "-", "np", "n/p"):
            return None
        try:
            return float(v.replace(",", "."))
        except ValueError:
            return None
    try:
        f = float(val)
        if f <= 0:
            return None
        return f
    except (TypeError, ValueError):
        return None


def parse_solde_report_amount(val: object) -> Optional[float]:
    """Convertit une cellule Solde Excel en montant_dû (>= 0) pour reports_membres."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    # Excel: solde négatif = retard à reporter
    if f < 0:
        return round(abs(f), 2)
    return 0.0


def month_columns_from_df(df: pd.DataFrame) -> dict[int, str]:
    norm_to_orig = {normalize_header(c): c for c in df.columns}
    mapping: dict[int, str] = {}
    for mi, key in enumerate(MONTH_ORDER_FR, start=1):
        if key in norm_to_orig:
            mapping[mi] = norm_to_orig[key]
    return mapping


def solde_columns_from_df(df: pd.DataFrame, sheet_year: int) -> list[tuple[int, str]]:
    """Colonnes solde Excel → (année d'exercice du report, nom de colonne).

    - « Solde 2025 » sur feuille Cotisations 2026 → report pour l'exercice 2026.
    - « Solde N-1 » → report pour l'année de la feuille (solde fin N-1 reprise en N).
    """
    out: list[tuple[int, str]] = []
    seen_years: set[int] = set()
    for c in df.columns:
        norm = normalize_header(c)
        m = SOLDE_YEAR_IN_COL_RE.search(norm)
        if m:
            report_annee = int(m.group(1)) + 1
        elif SOLDE_N1_COL_RE.match(norm):
            report_annee = sheet_year
        else:
            continue
        if report_annee in seen_years:
            continue
        seen_years.add(report_annee)
        out.append((report_annee, str(c)))
    return out


def resolve_col(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    norm_to_orig = {normalize_header(c): c for c in df.columns}
    for cand in candidates:
        nn = normalize_header(cand)
        if nn in norm_to_orig:
            return norm_to_orig[nn]
    return None


def cotisation_sheet_year(name: str) -> Optional[int]:
    m = COTISATIONS_SHEET_RE.match(name.strip())
    return int(m.group(1)) if m else None


def depenses_sheet_year(name: str) -> Optional[int]:
    m = DEPENSES_SHEET_RE.match(name.strip())
    return int(m.group(1)) if m else None


def last_day_of_month(year: int, month: int) -> date:
    last = calendar.monthrange(year, month)[1]
    return date(year, month, last)


def merge_member_field(prev: str, new_val: object) -> str:
    if new_val is None or (isinstance(new_val, float) and pd.isna(new_val)):
        return prev
    s = str(new_val).strip()
    if not s:
        return prev
    return s if not prev else prev


def load_existing_member_keys(conn: sqlite3.Connection) -> dict[tuple[str, str, str], int]:
    rows = conn.execute("SELECT id, nom, prenom, telephone FROM membres").fetchall()
    out: dict[tuple[str, str, str], int] = {}
    for r in rows:
        k = excel_member_key(r["nom"], r["prenom"], r["telephone"])
        out[k] = int(r["id"])
    return out


def contribution_note_import(sheet_label: str, month_fr: str) -> str:
    return f"Excel {sheet_label} {month_fr.capitalize()}"


def contribution_exists(
    conn: sqlite3.Connection,
    membre_id: int,
    date_iso: str,
    montant: float,
    note: str,
) -> bool:
    row = conn.execute(
        """
        SELECT 1 FROM contributions
        WHERE membre_id = ? AND date = ? AND ABS(montant - ?) < 0.001 AND note = ?
        LIMIT 1
        """,
        (membre_id, date_iso, montant, note),
    ).fetchone()
    return row is not None


def depense_exists(conn: sqlite3.Connection, description: str, date_iso: str, montant: float) -> bool:
    row = conn.execute(
        """
        SELECT 1 FROM depenses
        WHERE description = ? AND date = ? AND ABS(montant - ?) < 0.001
        LIMIT 1
        """,
        (description, date_iso, montant),
    ).fetchone()
    return row is not None


def insert_membre_from_import(
    conn: sqlite3.Connection,
    nom: str,
    prenom: str,
    telephone: str,
    village_origine: str,
    prefecture: str,
    email: str,
    adresse: str,
    date_inscription: date,
) -> int:
    cur = conn.execute(
        """
        INSERT INTO membres(reference, nom, prenom, telephone, village_origine, adresse, prefecture, email, date_inscription, actif)
        VALUES('', ?, ?, ?, ?, ?, ?, ?, ?, 1);
        """,
        (
            nom.strip(),
            prenom.strip(),
            telephone.strip(),
            village_origine.strip(),
            adresse.strip(),
            prefecture.strip(),
            email.strip(),
            to_iso(date_inscription),
        ),
    )
    new_id = int(cur.lastrowid)
    conn.execute("UPDATE membres SET reference = ? WHERE id = ?", (member_ref(new_id), new_id))
    log_activity(
        conn,
        type_operation="CREATE",
        entite="membre",
        entite_id=new_id,
        details="Import Excel "
        + fmt_member_compact(
            reference=member_ref(new_id),
            nom=nom.strip(),
            prenom=prenom.strip(),
            telephone=telephone,
            village=village_origine,
            prefecture=prefecture,
            email=email,
            adresse=adresse,
            date_inscription=to_iso(date_inscription),
        ),
    )
    return new_id


def parse_workbook_preview(
    raw: bytes,
    cotisation_sheets: list[str],
    depense_sheets: list[str],
    import_reports: bool,
    default_inscription: date,
) -> dict[str, Any]:
    xl = pd.ExcelFile(BytesIO(raw), engine="openpyxl")

    members: dict[MemberImportUid, dict[str, str]] = {}
    contributions: list[dict[str, object]] = []
    reports: list[dict[str, object]] = []
    depenses: list[dict[str, object]] = []

    month_labels = [
        "janvier",
        "fevrier",
        "mars",
        "avril",
        "mai",
        "juin",
        "juillet",
        "aout",
        "septembre",
        "octobre",
        "novembre",
        "decembre",
    ]

    for sheet in cotisation_sheets:
        sheet_year = cotisation_sheet_year(sheet)
        if sheet_year is None:
            continue
        df = pd.read_excel(BytesIO(raw), sheet_name=sheet, header=3, engine="openpyxl")
        col_nom = resolve_col(df, "Nom")
        col_prenom = resolve_col(df, "Prenom", "Prénom")
        col_tel = resolve_col(df, "telephone", "téléphone", "tel")
        col_pref = resolve_col(df, "Prefecture", "Préfecture", "prefecture")
        col_mail = resolve_col(df, "email", "e-mail")
        col_addr = resolve_col(df, "adresse")
        if not col_nom or not col_prenom:
            continue

        month_cols = month_columns_from_df(df)
        solde_cols = solde_columns_from_df(df, sheet_year) if import_reports else []

        for row_num, (_, row) in enumerate(df.iterrows(), start=1):
            nom_v = row[col_nom]
            prenom_v = row[col_prenom]
            if pd.isna(nom_v) or str(nom_v).strip() == "":
                continue
            if pd.isna(prenom_v) or str(prenom_v).strip() == "":
                continue

            uid = member_import_uid(nom_v, prenom_v, row[col_tel] if col_tel else "", sheet, row_num)
            tel_str = normalize_phone(row[col_tel]) if col_tel else ""
            pref = str(row[col_pref]).strip() if col_pref and pd.notna(row[col_pref]) else ""
            mail = str(row[col_mail]).strip() if col_mail and pd.notna(row[col_mail]) else ""
            addr = str(row[col_addr]).strip() if col_addr and pd.notna(row[col_addr]) else ""

            if uid not in members:
                members[uid] = {
                    "nom": str(nom_v).strip(),
                    "prenom": str(prenom_v).strip(),
                    "telephone": tel_str,
                    "village_origine": pref,
                    "prefecture": pref,
                    "email": mail,
                    "adresse": addr,
                }
            else:
                m = members[uid]
                m["telephone"] = merge_member_field(m["telephone"], tel_str if tel_str else None)
                m["email"] = merge_member_field(m["email"], mail if mail else None)
                m["adresse"] = merge_member_field(m["adresse"], addr if addr else None)
                m["village_origine"] = merge_member_field(m["village_origine"], pref if pref else None)
                m["prefecture"] = merge_member_field(m["prefecture"], pref if pref else None)

            for mi, col_name in month_cols.items():
                amt = parse_money_cell(row[col_name])
                if amt is None:
                    continue
                d = last_day_of_month(sheet_year, mi)
                contributions.append(
                    {
                        "member_uid": uid,
                        "nom": str(nom_v).strip(),
                        "prenom": str(prenom_v).strip(),
                        "telephone": tel_str,
                        "annee": sheet_year,
                        "mois": mi,
                        "montant": round(float(amt), 2),
                        "date_iso": d.isoformat(),
                        "note": contribution_note_import(sheet, month_labels[mi - 1]),
                    }
                )

            if import_reports:
                for report_annee, col_name in solde_cols:
                    rep_amt = parse_solde_report_amount(row[col_name])
                    if rep_amt is None:
                        continue
                    if rep_amt <= 0:
                        continue
                    reports.append(
                        {
                            "member_uid": uid,
                            "nom": str(nom_v).strip(),
                            "prenom": str(prenom_v).strip(),
                            "telephone": tel_str,
                            "report_annee": report_annee,
                            "montant_du": float(rep_amt),
                            "source_col": col_name,
                            "source_sheet": sheet,
                        }
                    )

    for sheet in depense_sheets:
        year_d = depenses_sheet_year(sheet)
        if year_d is None:
            continue
        df = pd.read_excel(BytesIO(raw), sheet_name=sheet, header=2, engine="openpyxl")
        col_lib = resolve_col(df, "Libellé", "Libelle")
        if not col_lib:
            continue
        month_cols = month_columns_from_df(df)

        for _, row in df.iterrows():
            lib_v = row[col_lib]
            if pd.isna(lib_v) or str(lib_v).strip() == "":
                continue
            desc = str(lib_v).strip()
            for mi, col_name in month_cols.items():
                amt = parse_money_cell(row[col_name])
                if amt is None:
                    continue
                d = last_day_of_month(year_d, mi)
                depenses.append(
                    {
                        "description": desc[:500],
                        "montant": round(float(amt), 2),
                        "date_iso": d.isoformat(),
                        "source_sheet": sheet,
                    }
                )

    return {
        "members": members,
        "contributions": contributions,
        "reports": reports,
        "depenses": depenses,
        "default_inscription": default_inscription,
        "xl_sheet_names": xl.sheet_names,
    }


def apply_import_bundle(conn: sqlite3.Connection, bundle: dict[str, Any]) -> dict[str, int]:
    """Écrit en base après validation. Retourne des compteurs."""
    members: dict[MemberImportUid, dict[str, str]] = bundle["members"]
    contributions: list[dict[str, object]] = bundle["contributions"]
    reports: list[dict[str, object]] = bundle["reports"]
    depenses_rows: list[dict[str, object]] = bundle["depenses"]
    default_inscription: date = bundle["default_inscription"]

    db_key_to_id = load_existing_member_keys(conn)
    uid_to_id: dict[MemberImportUid, int] = {}
    created = 0
    for uid, info in members.items():
        dk = excel_member_key(info["nom"], info["prenom"], info["telephone"])
        if dk in db_key_to_id:
            uid_to_id[uid] = db_key_to_id[dk]
            continue
        mid = insert_membre_from_import(
            conn,
            info["nom"],
            info["prenom"],
            info["telephone"],
            info["village_origine"],
            info["prefecture"],
            info["email"],
            info["adresse"],
            default_inscription,
        )
        db_key_to_id[dk] = mid
        uid_to_id[uid] = mid
        created += 1

    contrib_ins = 0
    contrib_skip = 0
    for c in contributions:
        uid = c["member_uid"]
        if uid not in uid_to_id:
            contrib_skip += 1
            continue
        mid = uid_to_id[uid]
        note = str(c["note"])
        montant = float(c["montant"])
        date_iso = str(c["date_iso"])
        if contribution_exists(conn, mid, date_iso, montant, note):
            contrib_skip += 1
            continue
        conn.execute(
            "INSERT INTO contributions(membre_id, montant, date, note) VALUES(?, ?, ?, ?)",
            (mid, montant, date_iso, note),
        )
        contrib_ins += 1

    dep_ins = 0
    dep_skip = 0
    for d in depenses_rows:
        desc = str(d["description"])
        montant = float(d["montant"])
        date_iso = str(d["date_iso"])
        if depense_exists(conn, desc, date_iso, montant):
            dep_skip += 1
            continue
        conn.execute(
            "INSERT INTO depenses(description, montant, date) VALUES(?, ?, ?)",
            (desc, montant, date_iso),
        )
        dep_ins += 1

    rep_ins = 0
    for r in reports:
        uid = r["member_uid"]
        if uid not in uid_to_id:
            continue
        mid = uid_to_id[uid]
        annee = int(r["report_annee"])
        amt = float(r["montant_du"])
        conn.execute(
            """
            INSERT INTO reports_membres(membre_id, annee, montant_du)
            VALUES(?, ?, ?)
            ON CONFLICT(membre_id, annee) DO UPDATE SET montant_du = excluded.montant_du;
            """,
            (mid, annee, amt),
        )
        rep_ins += 1

    log_activity(
        conn,
        type_operation="CREATE",
        entite="import_excel",
        entite_id=None,
        details=(
            f"Import Excel applique: membres_crees={created}, contributions={contrib_ins} "
            f"(ignores {contrib_skip}), depenses={dep_ins} (ignores {dep_skip}), reports={rep_ins}"
        ),
    )
    conn.commit()
    return {
        "membres_crees": created,
        "contributions": contrib_ins,
        "contributions_ignorees": contrib_skip,
        "depenses": dep_ins,
        "depenses_ignorees": dep_skip,
        "reports": rep_ins,
    }


def ensure_openpyxl_or_stop() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        st.error(
            "**openpyxl** est nécessaire pour lire les fichiers Excel (.xlsx).\n\n"
            "- **En local** : exécutez `python -m pip install openpyxl` "
            "(ou `python -m pip install -r requirements.txt` à la racine du projet).\n\n"
            "- **Sur Streamlit Community Cloud** : le fichier `requirements.txt` doit être "
            "à la **racine du dépôt Git** lié à l’app (avec une ligne `openpyxl>=3.1.0`), "
            "puis redémarrez l’app (**Manage app → Reboot**) ou refaites un déploiement."
        )
        st.stop()


def page_import_excel(conn: sqlite3.Connection) -> None:
    st.subheader("Import Excel")
    ensure_openpyxl_or_stop()
    render_page_guide(
        "Chargez votre classeur AGPM, vérifiez l'aperçu, complétez les téléphones si besoin, "
        "puis enregistrez en base. Aucune écriture sans validation."
    )

    st.session_state.setdefault("_import_bundle", None)

    with st.container(border=True):
        st.markdown("##### Étape 1 — Fichier")
        up = st.file_uploader("Classeur Excel (.xlsx)", type=["xlsx"], help="Ex. AGPM Association_2026.xlsx")

    if not up:
        st.info("📁 Déposez votre fichier Excel pour commencer.")
        return

    with st.container(border=True):
        st.markdown("##### Étape 2 — Options")
        default_date_ins = st.date_input(
            "Date d'inscription (nouveaux membres)",
            value=date(date.today().year, 1, 1),
            key="import_default_inscription",
        )
        import_reports = st.checkbox(
            "Importer les soldes N-1 / Solde YYYY (retards reportés)",
            value=True,
            help="Valeurs négatives Excel = retard converti en montant dû.",
        )

    raw = up.getvalue()
    try:
        xl_names = pd.ExcelFile(BytesIO(raw), engine="openpyxl").sheet_names
    except Exception as e:
        st.error(f"Lecture du fichier impossible : {e}")
        return

    cotisation_candidates = [n for n in xl_names if cotisation_sheet_year(n) is not None]
    depense_candidates = [n for n in xl_names if depenses_sheet_year(n) is not None]

    c_sel = st.multiselect(
        "Feuilles cotisations à importer",
        options=sorted(cotisation_candidates),
        default=sorted(cotisation_candidates),
    )
    d_sel = st.multiselect(
        "Feuilles dépenses à importer",
        options=sorted(depense_candidates),
        default=sorted(depense_candidates),
    )

    st.markdown("##### Étape 3 — Feuilles et aperçu")
    if st.button("Analyser le fichier et préparer l'aperçu", type="primary", use_container_width=True):
        try:
            bundle = parse_workbook_preview(
                raw,
                list(c_sel),
                list(d_sel),
                import_reports,
                default_date_ins,
            )
            st.session_state["_import_bundle"] = bundle
        except Exception as e:
            st.session_state["_import_bundle"] = None
            st.exception(e)

    bundle = st.session_state.get("_import_bundle")
    if not bundle:
        return

    existing = load_existing_member_keys(conn)
    members_dict: dict[MemberImportUid, dict[str, str]] = bundle["members"]
    rows_mem = []
    for uid, info in sorted(members_dict.items(), key=lambda x: (x[1]["nom"], x[1]["prenom"], str(x[0]))):
        dk = excel_member_key(info["nom"], info["prenom"], info["telephone"])
        rows_mem.append(
            {
                "_uid_json": uid_import_to_json(uid),
                "origine_ligne": member_import_row_label(uid),
                "nouveau": "Oui" if dk not in existing else "Non",
                "id_existant": str(existing[dk]) if dk in existing else "",
                "nom": info["nom"],
                "prenom": info["prenom"],
                "telephone": info["telephone"] if info["telephone"] else "",
                "email": info["email"],
                "prefecture": info["prefecture"],
            }
        )
    df_mem = pd.DataFrame(rows_mem)

    st.markdown("##### Étape 4 — Vérification des contacts")
    st.caption("Complétez les téléphones manquants avant l'enregistrement définitif.")
    st.caption(
        "Sans numéro dans Excel, chaque ligne de cotisation reste une entrée distincte (homonymes non fusionnés). "
        "Renseignez le téléphone ci-dessous puis fusionnez les doublons, avant d'écrire en base."
    )
    edited_mem = st.data_editor(
        df_mem,
        column_config={
            "_uid_json": st.column_config.TextColumn("UID", disabled=True, width="small"),
            "origine_ligne": st.column_config.TextColumn("Source Excel", disabled=True),
            "nouveau": st.column_config.TextColumn("Nouveau ?", disabled=True),
            "id_existant": st.column_config.TextColumn("Id base", disabled=True),
            "nom": st.column_config.TextColumn("Nom", disabled=True),
            "prenom": st.column_config.TextColumn("Prénom", disabled=True),
            "telephone": st.column_config.TextColumn("Téléphone", required=True),
            "email": st.column_config.TextColumn("Email", disabled=True),
            "prefecture": st.column_config.TextColumn("Préfecture", disabled=True),
        },
        hide_index=True,
        num_rows="fixed",
        use_container_width=True,
        key="import_editor_membres",
    )

    if st.button("Appliquer téléphones et fusionner les doublons", type="secondary"):
        sync_member_phones_from_editor(bundle, edited_mem)
        collapse_bundle_members_by_phone(bundle)
        st.session_state["_import_bundle"] = bundle
        st.rerun()
    st.caption(
        "Après saisie des numéros : utilisez ce bouton pour regrouper les fiches identiques "
        "(même nom, prénom et téléphone). L'enregistrement en base synchronise aussi les champs depuis le tableau."
    )

    cdf = pd.DataFrame(bundle["contributions"])
    if cdf.empty:
        st.warning("Aucune cotisation mensuelle numérique détectée dans les feuilles sélectionnées.")
    else:
        st.markdown(f"### Aperçu — contributions ({len(cdf)} lignes)")
        show_c = cdf.copy()

        def _deja_membre(r: pd.Series) -> bool:
            uid = r["member_uid"]
            info = bundle["members"].get(uid)
            if not info:
                return False
            dk = excel_member_key(info["nom"], info["prenom"], info["telephone"])
            return dk in existing

        show_c["telephone"] = show_c["member_uid"].apply(
            lambda u: bundle["members"].get(u, {}).get("telephone", "")
        )
        show_c["deja_membre"] = show_c.apply(_deja_membre, axis=1)
        show_c = show_c.drop(columns=["member_uid"], errors="ignore")
        st.dataframe(show_c.head(500), use_container_width=True)
        if len(show_c) > 500:
            st.caption(f"Affichage tronqué : {len(show_c) - 500} lignes supplémentaires non affichées.")

    ddf = pd.DataFrame(bundle["depenses"])
    if not ddf.empty:
        st.markdown(f"### Aperçu — dépenses ({len(ddf)} lignes)")
        st.dataframe(ddf.head(300), use_container_width=True)

    rdf = pd.DataFrame(bundle["reports"])
    if import_reports:
        if rdf.empty:
            st.info("Aucune valeur « Solde » retenue (vide ou zéro après conversion).")
        else:
            st.markdown(f"### Aperçu — reports membres ({len(rdf)} lignes)")
            st.dataframe(rdf.drop(columns=["member_uid"], errors="ignore").head(300), use_container_width=True)

    st.markdown("##### Étape 5 — Enregistrement")
    if st.button("Valider et enregistrer dans la base", type="primary", use_container_width=True):
        try:
            sync_member_phones_from_editor(bundle, edited_mem)
            collapse_bundle_members_by_phone(bundle)
            manques = validate_bundle_phones(bundle)
            if manques:
                st.error(
                    "Téléphone manquant pour les contacts suivants — complétez la colonne puis "
                    "« Appliquer téléphones… » ou réessayez :\n\n"
                    + "\n".join(f"- {m}" for m in manques[:40])
                    + (f"\n… ({len(manques) - 40} autres)" if len(manques) > 40 else "")
                )
                st.session_state["_import_bundle"] = bundle
            else:
                counts = apply_import_bundle(conn, bundle)
                st.success(
                    f"Import terminé — membres créés : {counts['membres_crees']}, "
                    f"cotisations : {counts['contributions']} (ignorées doublons : {counts['contributions_ignorees']}), "
                    f"dépenses : {counts['depenses']} (ignorées : {counts['depenses_ignorees']}), "
                    f"reports : {counts['reports']}."
                )
                st.session_state["_import_bundle"] = None
        except Exception as e:
            st.exception(e)


def get_app_password() -> Optional[str]:
    pwd = None
    try:
        pwd = st.secrets.get("app_password")
        if pwd is None:
            for value in st.secrets.values():
                if hasattr(value, "get"):
                    nested = value.get("app_password")
                    if nested:
                        pwd = nested
                        break
    except Exception:
        pwd = None
    if pwd is None:
        pwd = os.environ.get("AGPM_APP_PASSWORD")
    pwd = (str(pwd).strip() if pwd is not None else "")
    return pwd or None


def check_password(has_logo: bool) -> bool:
    if st.session_state.get("auth_ok"):
        return True

    expected = get_app_password()
    if not expected:
        st.error(
            "🔒 Accès non configuré. Définissez un mot de passe dans les **Secrets** "
            "de l'application (`app_password = \"…\"`) puis rechargez la page."
        )
        st.caption(
            "Streamlit Cloud : Manage app → Settings → Secrets. "
            "En local : `.streamlit/secrets.toml`."
        )
        return False

    col = st.columns([1, 2, 1])[1]
    with col:
        if has_logo:
            st.image(str(LOGO_PATH), use_container_width=True)
        st.markdown("### Accès réservé")
        st.caption("Cette application est protégée. Saisissez le mot de passe.")
        with st.form("login_form"):
            pwd = st.text_input("Mot de passe", type="password")
            if st.form_submit_button("Se connecter", type="primary", use_container_width=True):
                if pwd == expected:
                    st.session_state["auth_ok"] = True
                    st.rerun()
                else:
                    st.error("Mot de passe incorrect.")
    return False


def main() -> None:
    has_logo = LOGO_PATH.is_file()
    st.set_page_config(
        page_title="AGPM — Gestion association",
        page_icon=str(LOGO_PATH) if has_logo else "🏛️",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    render_app_styles()

    if not check_password(has_logo):
        st.stop()

    if has_logo:
        st.sidebar.image(str(LOGO_PATH), use_container_width=True)

    if st.sidebar.button("Se déconnecter", use_container_width=True):
        st.session_state.pop("auth_ok", None)
        st.rerun()

    render_storage_sidebar()
    conn = get_conn()

    monthly_amount = get_monthly_contribution(conn)
    st.title("AGPM")
    st.caption(
        "Association des Guinéens du Pays de Meaux · "
        f"Cotisation : {monthly_amount:.0f} EUR / mois"
    )

    menu = st.sidebar.radio(
        "Menu",
        [
            "📋 Tableau de bord",
            "👥 Membres",
            "💶 Cotisations",
            "📤 Dépenses",
            "📜 Journal",
            "📥 Import Excel",
        ],
    )
    st.sidebar.caption("Les montants et statuts sont calculés à partir des saisies réelles.")

    if menu == "👥 Membres":
        page_membres(conn)
    elif menu == "💶 Cotisations":
        page_contributions(conn)
    elif menu == "📤 Dépenses":
        page_depenses(conn)
    elif menu == "📜 Journal":
        page_activite(conn)
    elif menu == "📥 Import Excel":
        page_import_excel(conn)
    else:
        page_dashboard(conn)


if __name__ == "__main__":
    main()
